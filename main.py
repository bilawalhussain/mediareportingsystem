# flask imports
import io
import json
import os
from flask import Flask, render_template, request, jsonify, make_response,send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine
from werkzeug.security import generate_password_hash, check_password_hash
# imports for PyJWT authentication
import jwt
import numpy as np
from datetime import datetime, timedelta
from functools import wraps
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill,Font,colors
from openpyxl.styles import Alignment
from flask_cors import CORS, cross_origin
import pickle

# creates Flask object
app = Flask(__name__)
cors = CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'
app.config['CORS_HEADERS'] = 'x-access-token'
# configuration
# NEVER HARDCODE YOUR CONFIGURATION IN YOUR CODE
# INSTEAD CREATE A .env FILE AND STORE IN IT
app.config['SECRET_KEY'] = 'your secret key'
# database name
# app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root:@localhost/reporting_app'
# app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = True
# creates SQLALCHEMY object
# db = SQLAlchemy(app)

# blacklist = set()


# # Database ORMs
# class User(db.Model):
# 	id = db.Column(db.Integer, primary_key = True)
# 	username = db.Column(db.String(250), unique = True)
# 	email = db.Column(db.String(250))
# 	password = db.Column(db.String(250))

# class media_files(db.Model):
#     id = db.Column(db.Integer,primary_key = True)
#     filename = db.Column(db.String(250))
#     file = db.Column(db.LargeBinary)
#     username = db.Column(db.String(250))

#     def __init__(self,filename, file,username):
#         self.filename = filename
#         self.file = file
#         self.username = username
        
# class weekly_record(db.Model):
#     id = db.Column(db.Integer,primary_key = True)
#     week_start = db.Column(db.DateTime)
#     week_end = db.Column(db.DateTime)
#     impressions = db.Column(db.Integer)
#     clicks = db.Column(db.Integer)
#     trueview = db.Column(db.Integer)
#     Spendings = db.Column(db.String(250))
#     Percent_25_views = db.Column(db.Integer)
#     Percent_50_views = db.Column(db.Integer)
#     Percent_75_views = db.Column(db.Integer)
#     Percent_100_views = db.Column(db.Integer)
#     Budget_Spent = db.Column(db.String(250))
#     CTR = db.Column(db.String(250))
#     Daily_View_KPI = db.Column(db.Integer)
#     CPV_TrueView = db.Column(db.String(250))
#     CPV_Complete = db.Column(db.String(250))
#     Daily_KPI_Achievement = db.Column(db.String(250))
#     Three_Views = db.Column(db.Integer)
#     CPV_Three = db.Column(db.String(250))
#     file_id = db.Column(db.Integer)

#     def __init__(self,week_start,week_end,impressions,clicks,trueview,Spendings,Percent_25_views,Percent_50_views,Percent_75_views,Percent_100_views,Budget_Spent,CTR,Daily_View_KPI,CPV_TrueView,CPV_Complete,Daily_KPI_Achievement,Three_Views,CPV_Three,file_id):
#         self.week_start = week_start
#         self.week_end = week_end
#         self.impressions = impressions
#         self.clicks = clicks
#         self.trueview = trueview
#         self.Spendings = Spendings
#         self.Percent_25_views = Percent_25_views
#         self.Percent_50_views = Percent_50_views
#         self.Percent_75_views = Percent_75_views
#         self.Percent_100_views = Percent_100_views
#         self.Budget_Spent = Budget_Spent
#         self.CTR = CTR
#         self.Daily_View_KPI = Daily_View_KPI
#         self.CPV_TrueView = CPV_TrueView
#         self.CPV_Complete = CPV_Complete
#         self.Daily_KPI_Achievement = Daily_KPI_Achievement
#         self.Three_Views = Three_Views
#         self.CPV_Three = CPV_Three
#         self.file_id = file_id


# class annual_record(db.Model):
#     id = db.Column(db.Integer,primary_key = True)
#     day_start = db.Column(db.DateTime)
#     day_end = db.Column(db.DateTime)
#     impressions = db.Column(db.Integer)
#     clicks = db.Column(db.Integer)
#     trueview = db.Column(db.Integer)
#     Spendings = db.Column(db.String(250))
#     Percent_25_views = db.Column(db.Integer)
#     Percent_50_views = db.Column(db.Integer)
#     Percent_75_views = db.Column(db.Integer)
#     Percent_100_views = db.Column(db.Integer)
#     Budget_Spent = db.Column(db.String(250))
#     CTR = db.Column(db.String(250))
#     Daily_View_KPI = db.Column(db.Integer)
#     CPV_TrueView = db.Column(db.String(250))
#     CPV_Complete = db.Column(db.String(250))
#     Daily_KPI_Achievement = db.Column(db.String(250))
#     Three_Views = db.Column(db.Integer)
#     CPV_Three = db.Column(db.String(250))
#     file_id = db.Column(db.Integer)

#     def __init__(self,annual_total,file_id):
#         self.day_start = annual_total["day_start"]
#         self.day_end = annual_total["day_end"]
#         self.impressions = annual_total["impressions"]
#         self.clicks = annual_total["clicks"]
#         self.trueview = annual_total["trueview"]
#         self.Spendings = annual_total["spending"]
#         self.Percent_25_views = annual_total["25_percentage_view"]
#         self.Percent_50_views = annual_total["50_percentage_view"]
#         self.Percent_75_views = annual_total["75_percentage_view"]
#         self.Percent_100_views = annual_total["100_percentage_view"]
#         self.Budget_Spent = annual_total["budget_spent"]
#         self.CTR = annual_total["ctr"]
#         self.Daily_View_KPI = annual_total["weekly_view_kpi"]
#         self.CPV_TrueView = annual_total["cpv_trueview"]
#         self.CPV_Complete = annual_total["cpv_complete"]
#         self.Daily_KPI_Achievement = annual_total["weekly_kpi_achievement"]
#         self.Three_Views = annual_total["three_views"]
#         self.CPV_Three = annual_total["cpv_3_views"]
#         self.file_id = file_id

# class phases(db.Model):
#     id = db.Column(db.Integer,primary_key = True)
#     phase_name = db.Column(db.String(250))
    
# class phases_record(db.Model):
#     id = db.Column(db.Integer,primary_key = True)
#     phase_start = db.Column(db.DateTime)
#     phase_end = db.Column(db.DateTime)
#     impressions = db.Column(db.Integer)
#     clicks = db.Column(db.Integer)
#     trueview = db.Column(db.Integer)
#     Spendings = db.Column(db.String(250))
#     Percent_25_views = db.Column(db.Integer)
#     Percent_50_views = db.Column(db.Integer)
#     Percent_75_views = db.Column(db.Integer)
#     Percent_100_views = db.Column(db.Integer)
#     Budget_Spent = db.Column(db.String(250))
#     CTR = db.Column(db.String(250))
#     Daily_View_KPI = db.Column(db.Integer)
#     CPV_TrueView = db.Column(db.String(250))
#     CPV_Complete = db.Column(db.String(250))
#     Daily_KPI_Achievement = db.Column(db.String(250))
#     Three_Views = db.Column(db.Integer)
#     CPV_Three = db.Column(db.String(250))
#     phase_id = db.Column(db.Integer)
#     file_id = db.Column(db.Integer)

#     def __init__(self,phase_start,phase_end,impressions,clicks,trueview,Spendings,Percent_25_views,Percent_50_views,Percent_75_views,Percent_100_views,Budget_Spent,CTR,Daily_View_KPI,CPV_TrueView,CPV_Complete,Daily_KPI_Achievement,Three_Views,CPV_Three,phase_id,file_id):
#         self.phase_start = phase_start
#         self.phase_end = phase_end
#         self.impressions = impressions
#         self.clicks = clicks
#         self.trueview = trueview
#         self.Spendings = Spendings
#         self.Percent_25_views = Percent_25_views
#         self.Percent_50_views = Percent_50_views
#         self.Percent_75_views = Percent_75_views
#         self.Percent_100_views = Percent_100_views
#         self.Budget_Spent = Budget_Spent
#         self.CTR = CTR
#         self.Daily_View_KPI = Daily_View_KPI
#         self.CPV_TrueView = CPV_TrueView
#         self.CPV_Complete = CPV_Complete
#         self.Daily_KPI_Achievement = Daily_KPI_Achievement
#         self.Three_Views = Three_Views
#         self.CPV_Three = CPV_Three
#         self.phase_id = phase_id
#         self.file_id = file_id


# def check_if_token_in_blacklist(decrypted_token):
#     return decrypted_token in blacklist  
  
# # decorator for verifying the JWT
# def token_required(f):
# 	@wraps(f)
# 	def decorated(*args, **kwargs):
# 		token = None
# 		# jwt is passed in the request header
# 		if 'x-access-token' in request.headers:
# 			token = request.headers['x-access-token']
# 			# print(token)
# 		# return 401 if token is not passed
# 		if not token:
# 			return jsonify({'message' : 'Token is missing !!'}), 401
# 		if token in blacklist:
# 			return jsonify({
# 				'message' : 'Token is invalid'
# 			}), 401
# 		try:
# 			# decoding the payload to fetch the stored details
# 			data = jwt.decode(token, app.config['SECRET_KEY'],algorithms=['HS256'])
			
# 			# print("This is username",data["username"])
# 			current_user = User.query.filter_by(username = data['username']).first()
# 		except:
# 			return jsonify({
# 				'message' : 'Token is invalid'
# 			}), 401
# 		# returns the current logged in users contex to the routes
# 		return f(current_user, *args, **kwargs)

# 	return decorated


# # User Database Route
# # this route sends back list of users
# @app.route('/users', methods =['GET'])
# @cross_origin()
# @token_required
# def get_all_users(current_user):
# 	token = request.headers['x-access-token']
# 	if token in blacklist:
# 		return make_response(
# 			'Could not verify',
# 			401,
# 			{'WWW-Authenticate' : 'Basic realm ="Login required !!"'}
# 		)
# 	# print(current_user.username)
# 	# querying the database
# 	# for all the entries in it
# 	users = User.query.all()
# 	# converting the query objects
# 	# to list of jsons
# 	output = []
# 	for user in users:
# 		# appending the user data json
# 		# to the response list
# 		output.append({
# 			'username': user.username,
# 			'email' : user.email
# 		})

# 	return jsonify({'users': output})

# @app.route('/user', methods =['GET'])
# @cross_origin()
# @token_required
# def get_user(current_user):
# 	token = request.headers['x-access-token']
# 	if token in blacklist:
# 		return jsonify({'message':'Could not verify'},401)
	
# 	users = db.session.query(User).filter_by(username=current_user.username).first()
# 	db.session.close()
	
# 	if users is None:
# 		return jsonify({'message':'Could not verify'},401)

		
# 	return jsonify({
# 			'username': users.username,
# 			'email' : users.email
# 		},200)
# # route for logging user in
# @app.route('/login', methods =['POST'])
# @cross_origin()
# def login():
# 	# if 'x-access-token' in request.headers:
# 	# 	token = request.headers["x-access-token"]
# 	# 	if token in blacklist:
# 	# 		auth = request.form 
# 	# 		if not auth or not auth['username'] or not auth['password']:
# 	# 			# returns 401 if any email or / and password is missing
# 	# 			return make_response(
# 	# 				'Could not verify',
# 	# 				401,
# 	# 				{'WWW-Authenticate' : 'Basic realm ="Login required !!"'}
# 	# 			)

# 	# 		user = User.query.filter_by(username = auth['username']).first()

# 	# 		if not user:
# 	# 			# returns 401 if user does not exist
# 	# 			return make_response(
# 	# 				'Could not verify',
# 	# 				401,
# 	# 				{'WWW-Authenticate' : 'Basic realm ="User does not exist !!"'}
# 	# 			)

# 	# 		if check_password_hash(user.password, auth['password']):
# 	# 			# generates the JWT Token
# 	# 			token = jwt.encode({
# 	# 				'username': user.username,
# 	# 				'exp' : datetime.utcnow() + timedelta(minutes = 30)
# 	# 			}, app.config['SECRET_KEY'])

# 	# 			return make_response(jsonify({'token' : token}), 201)

# 	# 	return jsonify({
# 	# 		"message": "Already Logged in",
# 	# 		"token": token
# 	# 	},200)
# 	# # creates dictionary of form data
#     auth = request.json
    
#     ##print(auth)
#     # print(auth['username'])
#     # print(auth['password'])
#     if not auth or not auth['username'] or not auth['password']:
#     	# returns 401 if any email or / and password is missing
#         return make_response(
#     		'Could not verify',
#     		401,
#     		{'WWW-Authenticate' : 'Basic realm ="Login required !!"'}
#     	)

#     user = User.query.filter_by(username = auth['username']).first()
#     if not user:
#     	# returns 401 if user does not exist
#         return make_response(
#     		'username or password is incorrect',
#     		401,
#     		{'WWW-Authenticate' : 'Basic realm ="User does not exist !!"'}
#     	)
    
#     if check_password_hash(user.password, auth['password']):
#             # generates the JWT Token
#             token = jwt.encode({
#                 'username': user.username,
#                 'exp' : datetime.utcnow() + timedelta(minutes = 30)
#             }, app.config['SECRET_KEY'])
#             return make_response(jsonify({'token' : token}), 201)


    
#         # returns 403 if password is wrong
    
    
#     return make_response(
#     	'username or password is incorrect',
#     	403,
#     	{'WWW-Authenticate' : 'Basic realm ="Wrong Password !!"'}
#     )

# # signup route
# @app.route('/signup', methods =['POST'])
# @cross_origin()
# def signup():
# 	# creates a dictionary of the form data
# 	data = request.json

# 	# gets name, email and password
# 	username, email = data['username'], data['email']
# 	password = data['password']

# 	# checking for existing user
# 	user = User.query.filter_by(username = username).first()
# 	if not user:
# 		# database ORM object
# 		user = User(
# 			username = username,
# 			email = email,
# 			password = generate_password_hash(password)
# 		)
# 		# insert user
# 		db.session.add(user)
# 		db.session.commit()

# 		return make_response({'message':'Successfully registered.'}, 201)
# 	else:
# 		# returns 202 if user already exists
# 		return make_response({'message':'User already exists. Try a New one or Log in.'}, 202)

# # revoke current access token
# @app.route("/logout", methods=["GET"])
# @token_required
# @cross_origin()
# def logout_access_token(current_user):

	
# 	token = request.headers['x-access-token']

# 	if check_if_token_in_blacklist(token):
# 		return make_response(
# 			'Could not verify',
# 			401,
# 			{'WWW-Authenticate' : 'Basic realm ="Login required !!"'}
# 		)
# 	blacklist.add(token)
# 	return jsonify(message="Successfully logged out.")




# For reports ---------------------------------------




# # @app.route("/",methods=["GET","POST"])
# # @token_required
# # def index():
# #     db.create_all() # In case user table doesn't exists already. Else remove it.
# #     if "x-access-token" in blacklist:
# #         return redirect(url_for("login"))
# #     else:
# #         return render_template("home.html")


# def set_border_and_align(cell_info):
#         thin_border = Border(left=Side(style='thin'), 
#                           right=Side(style='thin'), 
#                           top=Side(style='thin'), 
#                           bottom=Side(style='thin'))

#         TEXT_ALIGNMENT = Alignment(horizontal='center',wrap_text=True)
#         cell_info.border = thin_border
#         cell_info.alignment = TEXT_ALIGNMENT
    
# def set_border_and_align_weekly(cell_info):
#         thin_border = Border(left=Side(style='thin'), 
#                           right=Side(style='thin'), 
#                           top=Side(style='thin'), 
#                           bottom=Side(style='thin'))

#         weekly_total_bg = "EBF1DE"

#         BG_PATTERN = PatternFill(start_color=weekly_total_bg,end_color=weekly_total_bg, fill_type = "solid")
            
#         TEXT_ALIGNMENT = Alignment(horizontal='center',wrap_text=True)
#         cell_info.border = thin_border
#         cell_info.alignment = TEXT_ALIGNMENT
#         cell_info.fill = BG_PATTERN

# def set_border_and_align_phase(cell_info):
#         thin_border = Border(left=Side(style='thin'), 
#                           right=Side(style='thin'), 
#                           top=Side(style='thin'), 
#                           bottom=Side(style='thin'))

#         phase_color = "FDE9D9"

#         PHASE_BG_PATTERN =  PatternFill(start_color=phase_color,end_color=phase_color, fill_type = "solid")
            
#         TEXT_ALIGNMENT = Alignment(horizontal='center',wrap_text=True)
#         cell_info.border = thin_border
#         cell_info.alignment = TEXT_ALIGNMENT
#         cell_info.fill = PHASE_BG_PATTERN

# def set_border_and_red_bg(cell_info):
#         thin_border = Border(left=Side(style='thin'), 
#                           right=Side(style='thin'), 
#                           top=Side(style='thin'), 
#                           bottom=Side(style='thin'))

#         phase_color = "8E1600"

#         PHASE_BG_PATTERN =  PatternFill(start_color=phase_color,end_color=phase_color, fill_type = "solid")
        
#         TEXT_ALIGNMENT = Alignment(horizontal='center',wrap_text=True)
#         cell_info.border = thin_border
#         cell_info.font = Font(color = colors.WHITE)
#         cell_info.alignment = TEXT_ALIGNMENT
#         cell_info.fill = PHASE_BG_PATTERN

# def set_border_and_black_bg(cell_info):
#         thin_border = Border(left=Side(style='thin'), 
#                           right=Side(style='thin'), 
#                           top=Side(style='thin'), 
#                           bottom=Side(style='thin'))

#         phase_color = "000000"

#         PHASE_BG_PATTERN =  PatternFill(start_color=phase_color,end_color=phase_color, fill_type = "solid")
        
#         TEXT_ALIGNMENT = Alignment(horizontal='center',wrap_text=True)
#         cell_info.border = thin_border
#         cell_info.font = Font(color = colors.WHITE)
#         cell_info.alignment = TEXT_ALIGNMENT
#         cell_info.fill = PHASE_BG_PATTERN

# def set_number_format(cell_info,type_of_format):

#     Formats = {
#         "percentage": '0%',
#         "comma_format": '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-',
#         "two_decimal_percentage": '0.00%',
#         "two_decimal_dollar":'_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-',
#         "five_digit":'#,##0',
#         "four_decimal_dollar":'_("$"* #,##0.0000_);_("$"* (#,##0.0000);_("$"* "-"????_);_(@_)'
        
#     }
#     cell_info.number_format = Formats[type_of_format]

# def check_filename_already_exist(filename,token_user):



#     res = db.session.query(media_files).filter_by(filename = filename, username = token_user.username).first()
#     db.session.close()
#     if res is None:
#         return False
#     return True

# @app.route("/reports",methods=["POST"])
# @cross_origin()
# @token_required

# def youtube(current_user):
#         if request.method=="POST":
#             filename = request.form["filename"]
#             if check_filename_already_exist(filename,current_user):
#                 return jsonify({
#                 "Message" : "Report name already exist. Try a new one",
#                 "status" : 400
#             },400)
#             phase_1_deadline = request.form["phase_1"]
#             phase_1_deadline = phase_1_deadline.replace("-","/")

#             phase_2_deadline = request.form['phase_2']
#             phase_2_deadline = phase_2_deadline.replace("-","/")

#             phase_1_budget_org = float(request.form['phase_1_budget'])
#             phase_1_budget = float(request.form['phase_1_budget'])
#             phase_2_budget = float(request.form['phase_2_budget'])
#             phase_2_budget_org = float(request.form['phase_2_budget'])

#             try:
#                 monthly_view_kpi_w1_org = float(request.form["monthly_view_kpi_w1"])
#                 monthly_view_kpi_w1 = float(request.form["monthly_view_kpi_w1"])
#                 monthly_view_kpi_w2 = float(request.form["monthly_view_kpi_w2"])
#                 monthly_view_kpi_w2_org = float(request.form["monthly_view_kpi_w2"])
#             except:
#                 monthly_view_kpi_w1_org = 0.0
#                 monthly_view_kpi_w1 = 0.0
#                 monthly_view_kpi_w2 = 0.0
#                 monthly_view_kpi_w2_org = 0.0
#             monthly_visit_kpi_w1 = request.form["monthly_visit_kpi_w1"]
#             monthly_visit_kpi_w2 = request.form["monthly_visit_kpi_w2"]



#             file_youtube_trueview = request.files["file"]

#             df = pd.read_excel(file_youtube_trueview)
           
#             # Grab all the data 


#             column_values = df.columns.ravel()
#             # These are the columns of the excel sheet data 
#             # We have all the columns of the file 

           
#             # Spending column checks 
  

#             df = pd.pivot_table(df, index=['Date'],values=column_values,aggfunc='sum')
#             # Getting all the columns data according to the file 


#             # Sum is coming from here
           
#             # we have 2D Array with the data 
            
#             weekly_list = []
#             # Weeks data is stored in this list 

#             j=0
#             # Startig the index from 0 and goes according to week 


#             while j < len(df):
#                 total_impressions = 0
#                 # In this we store total impressions 

#                 total_clicks = 0
#                 # In this we store total clicks 

#                 total_three_views = 0
#                 # In this we store three views  
                
#                 spending = 0
#                 # In this we store spendings 

#                 total_3_views = 0
#                 # In this we store total true view 


#                 total_trueview = 0
#                 # In this we store total true view 
                
#                 percent_25 = 0
#                 # In this we store 25 % Views 


#                 percent_50 = 0
#                 # In this we store 50% views 
                

#                 percent_75 = 0
#                 # In this we store 75% Views 


#                 percent_100 = 0
#                 # In this we store 100% Views 


#                 if j+7 < len(df):
#                 # If the 7 days index is less than the len of total records 
#                     end_date = df.index[j+7]
#                     end_index=j+7
#                     # Then store the end_date with j+7
#                 else:
#                 # Else 
#                     end_date = df.index[len(df)-1]
#                     end_index = len(df)
#                     # Store end date with length of df - 1 and end_index to length of df .

                
#                 data ={
#                     "date_start": df.index[j],
#                     "date_end": end_date,
#                     "start_index": j,
#                     "end_index": end_index
#                 }
#                 # Store the data in the data obj with all the rows record 
               

#                 for i in range(j,end_index):
#                 # Looping through each week's data and totaling impressions and clicks  
#                     total_impressions += df["Impressions"][i]

#                     total_clicks += df["Clicks"][i]
#                     spending += df["Spendings"][i]
#                     try:
#                         total_three_views += df["Three Views"][i]
#                     except:
#                         pass
#                     try:       
#                         total_trueview += df["Trueviews"][i]
#                         percent_25 += df["25% Views"][i]
#                         percent_50 += df["50% Views"][i]
#                         percent_75 += df["75% Views"][i]
#                         percent_100 += df["100% Views"][i]
#                     except:
#                         pass
                

                

#                 data["impressions"] = total_impressions
#                 data["clicks"] = total_clicks
#                 data["spending"] = spending

#                 data["three_views"] = total_three_views
#                 data["trueview"] = total_trueview
#                 data["percent_25"] = percent_25
#                 data["percent_50"] = percent_50
#                 data["percent_75"] = percent_75
#                 data["percent_100"] = percent_100

#                 weekly_list.append(data)
#                 j=j+7
                   
#             ##print(weekly_list)


#             # print(weekly_list)
#             ending_index_phase_1=0
#             start_index_phase_2 = 0
#             for i in range(0,len(df)):
#                 # print(df.index[i]," : " ,phase_1_deadline)
#                 if df.index[i] == phase_1_deadline:
#                     ending_index_phase_1 = i
#                     start_index_phase_2 = i+1
#                     break
            
#             # Phases ------------------------------
         
#             total_impressions = 0
#             total_clicks = 0
#             spending = 0
#             total_three_views = 0 
#             total_trueview = 0
#             percent_25 = 0
#             percent_50 = 0
#             percent_75 = 0
#             percent_100 = 0

            
#             end_date = df.index[ending_index_phase_1]
#             end_index = ending_index_phase_1+1
#             data ={
#                 "date_start": df.index[0],
#                 "date_end": end_date,
#                 "start_index": 0,
#                 "end_index": end_index
#             }
#             phase_1 = 0  
#             # print(data)
#             for i in range(0,end_index):
#                 total_impressions += df["Impressions"][i]
#                 total_clicks += df["Clicks"][i]
#                 spending += df["Spendings"][i]
              
#                 try:
#                     total_three_views += df["Three Views"][i]
#                 except:
#                     pass

#                 try:
#                     total_trueview += df["Trueviews"][i]
#                     percent_25 += df["25% Views"][i]
#                     percent_50 += df["50% Views"][i]
#                     percent_75 += df["75% Views"][i]
#                     percent_100 += df["100% Views"][i]
#                 except:
#                     pass
#             data["impressions"] = total_impressions
#             data["clicks"] = total_clicks
#             data["spending"] = spending
#             data["three_views"] = total_three_views
#             data["trueview"] = total_trueview
#             data["percent_25"] = percent_25
#             data["percent_50"] = percent_50
#             data["percent_75"] = percent_75
#             data["percent_100"] = percent_100
            
#             phase_1 = data
#             # print(phase_1)
#             # --------------------------- Phase 1 Ends Here ---------------------------------------------

#             # ---------------------------Phase 2 starts here --------------------------------------------
#             total_impressions = 0  # Total Impressions for phase 2 
#             total_clicks = 0        # Total Clicks for phase 2 
#             spending = 0            # Total spendings of phase 2 
#             total_three_views = 0
#             total_trueview = 0      # Total trueview for phase 2 
#             percent_25 = 0          # Total 25% for phase 2 
#             percent_50 = 0          # Total 50% for phase 2 
#             percent_75 = 0          # Total 75%  of phase 2 
#             percent_100 = 0         # Total 100% of phase 2 

#             start_date_phase_2 =df.index[start_index_phase_2]
            
#             # print(start_index_phase_2)
#             end_date = df.index[len(df)-1]        # Storing the ending date of phase 2 
#             end_index = len(df)                 # Storing ending index of phase 2 
#             # print(end_index)
#             data ={                             # We will store data in this object including : start date and end date with start index and end index 
#                 "date_start": start_date_phase_2,      # Start date would be ending_date of phase 1 + 1  index
#                 "date_end": end_date,                   # ending date of phase 2 
#                 "start_index": start_index_phase_2,        # start index of phase 2 
#                 "end_index": end_index                      # Ending index of phase 2 
#             }
#             phase_2 = 0                                     # This will store the data of the phase 2                               
#             # print(data)
#             for i in range(start_index_phase_2,end_index):      # Looping through the starting index phase 2 and ending index phase 2 
#                 total_impressions += df["Impressions"][i]           # Total impressions are added in the previous total impressions variable 
#                 total_clicks += df["Clicks"][i]                     # Total clicks are added in the previous total clicks variable
#                 spending += df["Spendings"][i]             # Total Spendings are added in the previous Spendings variable
#                 try:
#                     total_three_views += df["Three Views"][i]          # Total trueview are added in the previous total trueview variable
#                 except:
#                     pass

#                 try:
#                     total_trueview += df["Trueviews"][i]          # Total trueview are added in the previous total trueview variable
#                     percent_25 += df["25% Views"][i]     # Total 25 percentage are added in the previous 25 percentage variable 
#                     percent_50 += df["50% Views"][i]           # Total 50 percentage are added in the previous 50 percentage variable
#                     percent_75 += df["75% Views"][i]     # Total 75 percentage are added in the previous 75 percentage variable
#                     percent_100 += df["100% Views"][i]          # Total 100 percentage are added in the previous 100 percentage variable
#                 except:
#                     pass

#             data["impressions"] = total_impressions
#             data["clicks"] = total_clicks
#             data["spending"] = spending

#             data["three_views"] = total_three_views
#             data["trueview"] = total_trueview
#             data["percent_25"] = percent_25
#             data["percent_50"] = percent_50
#             data["percent_75"] = percent_75
#             data["percent_100"] = percent_100

#             phase_2 = data 
#             # print(phase_2)
#             #-------------------- Phase 2 ends here --------------------------------
#             annual_total = {
#                     "impressions": 0.0,
#                     "clicks": 0.0,
#                     "three_views": 0.0,
#                     "trueview": 0.0,
#                     "25_percentage_view": 0.0,
#                     "50_percentage_view": 0.0,
#                     "75_percentage_view": 0.0,
#                     "100_percentage_view": 0.0,
#                     "spending": 0.0,
#                     "budget_spent": 0.0,
#                     "ctr": 0.0,
#                     "weekly_view_kpi": 0.0,
#                     "cpv_3_views": 0.0,
#                     "cpv_trueview": 0.0,
#                     "cpv_complete": 0.0,
#                     "weekly_kpi_achievement": 0.0

#             }
           
#             # Do Calculations 

#             # So now we have store all the data in the backend now we have to store it in excel 
#             # Load template 
            
#             default_template = load_workbook("static/excel_template/youtube_template.xlsx")

#             # go to specific tab of the excel sheet 
#             youtube_trueview_sheet = default_template.active
            
            
            

#             labels_names = {
#                 "LABEL_DATE" : "Date",
#                 "LABEL_IMPRESSIONS": "Impressions",
#                 "LABEL_CLICKS": "Clicks",
#                 "LABEL_THREE_VIEWS": "3 Views",
#                 "LABEL_TRUEVIEW":"TrueView",
#                 "LABEL_25_PERCENTAGE_VIEW":"25% View of Video",
#                 "LABEL_50_PERCENTAGE_VIEW":"50% View of Video",
#                 "LABEL_75_PERCENTAGE_VIEW":"75% View of Video",
#                 "LABEL_100_PERCENTAGE_VIEW":"100% View",
#                 "LABEL_SPENDINGS":"Spending",
#                 "LABEL_BUDGET_SPENT":"Budget Spent",
#                 "LABEL_CTR":"CTR",
#                 "LABEL_DAILY_VIEW_KPI":"Daily View KPI",
#                 "LABEL_CPV_3_VIEWS":"CPV (3\'s)",
#                 "LABEL_CPV_TRUEVIEW":"CPV (Trueview)",
#                 "LABEL_CPV_COMPLETE":"CPV (Complete)",
#                 "LABEL_DAILY_KPI_ACHIEVEMENT":"Daily KPI Achievement"
#             }
#             columns_location ={
#                 "Date" : 1,
#                 "Impressions": 2,
#                 "Clicks": 3,
#                 "Three Views": 4,
#                 "Trueview": 5,
#                 "25_percent": 6,
#                 "50_percent": 7,
#                 "75_percent": 8,
#                 "100_percent": 9,
#                 "spendings": 10,
#                 "budget_spent": 11,
#                 "ctr": 12,
#                 "daily_view_kpi": 13,
#                 "cpv_3_views": 14,
#                 "cpv_trueview": 15,
#                 "cpv_complete": 16,
#                 "daily_kpi_achievement": 17,
                
#                 # W1 Views  - Will also be used for budget phase 1 and phase 2
#                 "w1_heading_start": 19,
#                 "w1_heading_end": 21,
#                 "w1_views_heading": 19,
#                 "w1_monthly_view_heading": 20,
#                 "w1_kpi_achievement_heading": 21,

#                 "w1_views_value": 19,
#                 "w1_monthly_view_value": 20,
#                 "w1_kpi_achievement_value": 21,
                
#                 # W2 Views 
#                 "w2_heading_start": 23,
#                 "w2_heading_end": 25,
#                 "w2_views_heading": 23,
#                 "w2_monthly_view_heading": 24,
#                 "w2_kpi_achievement_heading": 25,

#                 "w2_views_value": 23,
#                 "w2_monthly_view_value": 24,
#                 "w2_kpi_achievement_value": 25,

#             }
#             columns_label  = {
#                 labels_names["LABEL_DATE"]: columns_location["Date"] ,
#                 # This is the date column location 
#                 labels_names["LABEL_IMPRESSIONS"]:columns_location["Impressions"],
#                 # This is the date Impressions location 
#                 labels_names["LABEL_CLICKS"]:columns_location["Clicks"],
#                 # This is the date Clicks location 
#                 labels_names["LABEL_THREE_VIEWS"]:columns_location["Three Views"],
#                 # This is the date Three Views location 
#                 labels_names["LABEL_TRUEVIEW"]:columns_location["Trueview"],
#                 # This is the date Trueview location 
#                 labels_names["LABEL_25_PERCENTAGE_VIEW"]:columns_location["25_percent"],
#                 # This is the date 25_percent location 
#                 labels_names["LABEL_50_PERCENTAGE_VIEW"]:columns_location["50_percent"],
#                 # This is the date 50_percent location 
#                 labels_names["LABEL_75_PERCENTAGE_VIEW"]:columns_location["75_percent"],
#                 # This is the date 75_percent location 
#                 labels_names["LABEL_100_PERCENTAGE_VIEW"]:columns_location["100_percent"],
#                 # This is the date 100_percent location 
#                 labels_names["LABEL_SPENDINGS"]:columns_location["spendings"],
#                 # This is the date spendings location 
#                 labels_names["LABEL_BUDGET_SPENT"]:columns_location["budget_spent"],
#                 # This is the date budget_spent location 
#                 labels_names["LABEL_CTR"]: columns_location["ctr"],
#                 # This is the date ctr location 
#                 labels_names["LABEL_DAILY_VIEW_KPI"]: columns_location["daily_view_kpi"],
#                 # This is the date daily_view_kpi location 
#                 labels_names["LABEL_CPV_3_VIEWS"]:columns_location["cpv_3_views"],
#                 # This is the cpv_3_views location 
#                 labels_names["LABEL_CPV_TRUEVIEW"]:columns_location["cpv_trueview"],
#                 # This is the cpv_trueview column location 
#                 labels_names["LABEL_CPV_COMPLETE"]:columns_location["cpv_complete"],
#                 # This is the cpv_complete column location 
#                 labels_names["LABEL_DAILY_KPI_ACHIEVEMENT"]:columns_location["daily_kpi_achievement"]
#             }
#             # Here we are storing column locations 
           
            

#             # labels with column name 
#             thin_border = Border(left=Side(style='thin'), 
#                           right=Side(style='thin'), 
#                           top=Side(style='thin'), 
#                           bottom=Side(style='thin'))


            
#             TEXT_ALIGNMENT = Alignment(horizontal='center')


#             # Table with weekly data 

#             for key,value in columns_label.items():
#                 cell_location = youtube_trueview_sheet.cell(row=1, column=value)
#                 cell_location.value =  key    # This is the title of the table with title : Date
#                 cell_location.border = thin_border
#                 cell_location.alignment = TEXT_ALIGNMENT
#                 cell_location.fill = PatternFill(start_color="DAEEF3",end_color="DAEEF3", fill_type = "solid")
            

            



#             # Printing column names in excel sheet     
                
#             weekly_counter = 0              # In the start weekly counter will be 0 
#             weekly_number = 0               # Weekly number will be 0 
#             row_num = 2                     # Data storing location by row
#             previous_budget = 0  
#             total_weekly_kpi = 0 
#             previous_budget_by_week = 0
#             total_phase_1_daily_kpi= 0

#             # First thing is to run a loop to store the total data  
#             for i in range(2,len(df)+2):            # Loop to start adding in the file 
           
                
#                     previous_budget = previous_budget+ df["Spendings"][i-2]
      
#                     total_weekly_kpi = total_weekly_kpi + (monthly_view_kpi_w1 / 30)
#                     total_phase_1_daily_kpi = total_phase_1_daily_kpi + monthly_view_kpi_w1 / 30 
                    

#                     weekly_counter +=1              # counting the values and checking if the records are exceeding and getting more than week then print week 
#                     if weekly_counter >=7 or i == len(df)+1:          # if the weekly counter is greater than or equal to 7 then print week 
#                         if weekly_number == 0:
#                             annual_total["day_start"] = str(weekly_list[weekly_number]["date_start"])
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=1)
#                         cell_info.value =  str(weekly_number+1)+ "W\n" + str(weekly_list[weekly_number]["date_start"]) +" - " + str(weekly_list[weekly_number]["date_end"]) 
#                         set_border_and_align(cell_info)
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Impressions"])
#                         cell_info.value = weekly_list[weekly_number]['impressions']      # Printing impressions for week
#                         annual_total["impressions"]+= weekly_list[weekly_number]['impressions']
                        
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"comma_format")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Clicks"])
#                         cell_info.value = weekly_list[weekly_number]['clicks']  # Printing clicks for week
#                         annual_total["clicks"]+= weekly_list[weekly_number]['clicks']
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"comma_format")
                        
#                         try:
#                             cell_info = youtube_trueview_sheet.cell(row=row_num,column=columns_location["Three Views"])
#                             cell_info.value = weekly_list[weekly_number]["three_views"]
#                             annual_total["three_views"] += weekly_list[weekly_number]["three_views"]
#                             set_border_and_align(cell_info)
#                             set_number_format(cell_info,"comma_format")
#                         except:
#                             pass

#                         try:
                            

#                             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Trueview"])
#                             cell_info.value = weekly_list[weekly_number]['trueview'] # Printing trueview for week
#                             annual_total["trueview"]+= weekly_list[weekly_number]['trueview']
#                             set_border_and_align(cell_info)
#                             set_number_format(cell_info,"comma_format")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["25_percent"])
#                             cell_info.value = weekly_list[weekly_number]['percent_25']  # Printing 25% views for week
#                             annual_total["25_percentage_view"]+= weekly_list[weekly_number]['percent_25']
#                             set_border_and_align(cell_info)
#                             set_number_format(cell_info,"comma_format")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["50_percent"])
#                             cell_info.value = weekly_list[weekly_number]['percent_50'] # Printing 50% views for week
#                             annual_total["50_percentage_view"]+= weekly_list[weekly_number]['percent_50']
#                             set_border_and_align(cell_info)
#                             set_number_format(cell_info,"comma_format")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["75_percent"])
#                             cell_info.value = weekly_list[weekly_number]['percent_75']    # Printing 75% for week
#                             annual_total["75_percentage_view"]+= weekly_list[weekly_number]['percent_75']
#                             set_border_and_align(cell_info)
#                             set_number_format(cell_info,"comma_format")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["100_percent"])
#                             cell_info.value = weekly_list[weekly_number]['percent_100']       # Printing 100% for week
#                             annual_total["100_percentage_view"]+= weekly_list[weekly_number]['percent_100']
#                             set_border_and_align(cell_info)
#                             set_number_format(cell_info,"comma_format")

#                         except:
#                             pass    
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["spendings"])
#                         cell_info.value = weekly_list[weekly_number]['spending']      # Printing spendings for week
#                         annual_total["spending"]+= weekly_list[weekly_number]['spending']
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"two_decimal_dollar")

                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["budget_spent"])
#                         cell_info.value = (previous_budget_by_week + weekly_list[weekly_number]['spending'])  / phase_1_budget                # This is the Budget Spent 
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"percentage")
                                                
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["ctr"])
#                         cell_info.value = weekly_list[weekly_number]['clicks'] / weekly_list[weekly_number]['impressions']                # This is the CTR 
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"two_decimal_percentage")
                        
#                         try:
#                             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["cpv_3_views"])
#                             cell_info.value = weekly_list[weekly_number]['spending']   / weekly_list[weekly_number]['three_views'] 
#                             set_border_and_align(cell_info)
#                             set_number_format(cell_info,"four_decimal_dollar")

#                         except:
#                             pass

#                         try:

#                             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["daily_view_kpi"])
#                             cell_info.value = total_weekly_kpi                   # This is the Daily View KPI 
#                             annual_total["weekly_view_kpi"] = annual_total["weekly_view_kpi"] + total_weekly_kpi
#                             set_border_and_align(cell_info)
#                             set_number_format(cell_info,"five_digit")
                            


#                             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["cpv_trueview"])
#                             cell_info.value = weekly_list[weekly_number]['percent_100']   / weekly_list[weekly_number]['trueview'] 
#                             set_border_and_align(cell_info)
#                             set_number_format(cell_info,"four_decimal_dollar")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["cpv_complete"])
#                             cell_info.value =   weekly_list[weekly_number]['spending']  / weekly_list[weekly_number]['percent_100'] 
#                             set_border_and_align(cell_info)
#                             set_number_format(cell_info,"four_decimal_dollar")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["daily_kpi_achievement"])
#                             cell_info.value =   weekly_list[weekly_number]['percent_100']  / total_weekly_kpi
#                             set_border_and_align(cell_info)
#                             set_number_format(cell_info,"percentage")
#                         except:
#                             pass
#                         # print(weekly_list[weekly_number]['percent_100']," : ",total_weekly_kpi)


#                         previous_budget_by_week=previous_budget_by_week + weekly_list[weekly_number]['spending']
    
                        
#                         total_weekly_kpi=  0
#                         weekly_number +=1           # incrementing week 
#                         row_num+=1                  # Incrementing the row so we can store afterwards record 
#                         weekly_counter = 0          # making the counter back to 0 
                    
#                     # Now we are ready to deal with phase 1 and phase 2 
#                     # So if the wee


            


#             # Show annual total here ------------------
#             annual_total["day_end"] =  str(weekly_list[weekly_number-1]["date_end"]) 
#             annual_total["budget_spent"] = annual_total["spending"]/(float(phase_1_budget_org) + float(phase_2_budget_org))
            
#             annual_total["ctr"] = float(annual_total["clicks"])/ float(annual_total["impressions"])


#             try:
#                 annual_total["cpv_trueview"] = float(annual_total["spending"])/ float(annual_total["three_views"])
#             except:
#                 pass
            
#             try:
#                 annual_total["cpv_trueview"] = float(annual_total["spending"])/ float(annual_total["trueview"])
                
#                 annual_total["cpv_complete"] = float(annual_total["spending"])/ float(annual_total["100_percentage_view"])
                
#                 annual_total["weekly_kpi_achievement"] = float(annual_total["100_percentage_view"])/ float(annual_total["weekly_view_kpi"])
#             except:
#                 pass
#             cell_info = youtube_trueview_sheet.cell(row=row_num, column=1)
#             cell_info.value = "Total"             # Printing weekly Total
#             set_border_and_red_bg(cell_info)
            
#             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Impressions"])
#             cell_info.value = annual_total['impressions']      # Printing impressions for week
#             set_border_and_red_bg(cell_info)
#             set_number_format(cell_info,"comma_format")
            
#             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Clicks"])
#             cell_info.value = annual_total['clicks']  # Printing clicks for week
#             set_border_and_red_bg(cell_info)
#             set_number_format(cell_info,"comma_format")

#             try:
#                 cell_info = youtube_trueview_sheet.cell(row=row_num,column=columns_location["Three Views"])
#                 cell_info.value = annual_total["three_views"]
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"comma_format")

#             except:
#                 pass


#             try:
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Trueview"])
#                 cell_info.value = annual_total['trueview'] # Printing trueview for week
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"comma_format")
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["25_percent"])
#                 cell_info.value = annual_total['25_percentage_view']  # Printing 25% views for week
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"comma_format")
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["50_percent"])
#                 cell_info.value = annual_total['50_percentage_view'] # Printing 50% views for week
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"comma_format")
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["75_percent"])
#                 cell_info.value = annual_total['75_percentage_view']    # Printing 75% for week
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"comma_format")
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["100_percent"])
#                 cell_info.value = annual_total['100_percentage_view']       # Printing 100% for week
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"comma_format")
#             except:
#                 pass

#             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["spendings"])
#             cell_info.value = annual_total['spending']      # Printing spendings for week
#             set_border_and_red_bg(cell_info)
#             set_number_format(cell_info,"two_decimal_dollar")
            
#             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["budget_spent"])
#             cell_info.value = annual_total["budget_spent"]      # Printing spendings for week
#             set_border_and_red_bg(cell_info)
#             set_number_format(cell_info,"percentage")
            
#             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["ctr"])
#             cell_info.value = annual_total["ctr"]       # Printing spendings for week
#             set_border_and_red_bg(cell_info)
#             set_number_format(cell_info,"two_decimal_percentage")
            
#             try:
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["cpv_3_views"])
#                 cell_info.value = annual_total["cpv_3_views"]               # This is the spending 
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"four_decimal_dollar")
#             except:
#                 pass

#             try:
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["daily_view_kpi"])
#                 cell_info.value = annual_total["weekly_view_kpi"]                  # This is the spending 
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"five_digit")
                
                
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["cpv_trueview"])
#                 cell_info.value = annual_total["cpv_trueview"]               # This is the spending 
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"four_decimal_dollar")
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["cpv_complete"])
#                 cell_info.value = annual_total["cpv_complete"] 
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"four_decimal_dollar")
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["daily_kpi_achievement"])
#                 cell_info.value = annual_total["weekly_kpi_achievement"]           # This is the spending 
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"percentage")
#             except:
#                 pass

#             #------------------------------------------
#             starting_row_Summary = 1 

#             # W1 starting below this ---------------------------------
#             # Showing monthly view kpi and monthly visit kpi 

#             # View starts -------------------------------------------
#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary,column=columns_location["w1_heading_start"])
#             cell_info.value = "W1 KPI Summary"
#             youtube_trueview_sheet. merge_cells(start_row=starting_row_Summary, start_column=columns_location["w1_heading_start"], end_row=starting_row_Summary, end_column=columns_location["w1_heading_end"])
#             set_border_and_red_bg(cell_info)
#             try:
#                 cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+1,column=columns_location["w1_views_heading"])
#                 cell_info.value = "Views"
#                 set_border_and_black_bg(cell_info)
                
#                 cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+1,column=columns_location["w1_monthly_view_heading"])
#                 cell_info.value = "Monthly View KPI "
#                 set_border_and_black_bg(cell_info)

#                 cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+1,column=columns_location["w1_kpi_achievement_heading"])
#                 cell_info.value = "KPI Achievement"
#                 set_border_and_black_bg(cell_info)

#                 cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+2,column=columns_location["w1_views_value"])
#                 cell_info.value = phase_1["percent_100"]
#                 set_border_and_align(cell_info)

#                 set_number_format(cell_info,"five_digit")
#                 cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+2,column=columns_location["w1_monthly_view_value"])
#                 cell_info.value = monthly_view_kpi_w1_org
#                 set_border_and_align(cell_info)

#                 set_number_format(cell_info,"five_digit")
#                 cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+2,column=columns_location["w1_kpi_achievement_value"])
#                 cell_info.value = float(phase_1["percent_100"])/float(monthly_view_kpi_w1_org)
#                 set_border_and_align(cell_info)

#                 set_number_format(cell_info,"percentage")
#             except:
#                 pass

#             # View Ends here -----------------------------------------

#             # From here visits starts ---------------------------------


#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+3,column=columns_location["w1_views_heading"])
#             cell_info.value = "Visits"
#             set_border_and_black_bg(cell_info)
            
#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+3,column=columns_location["w1_monthly_view_heading"])
#             cell_info.value = "Monthly Visit KPI"
#             set_border_and_black_bg(cell_info)

#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+3,column=columns_location["w1_kpi_achievement_heading"])
#             cell_info.value = "KPI Achievement"
#             set_border_and_black_bg(cell_info)

#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+4,column=columns_location["w1_views_value"])
#             # cell_info.value = phase_1["percent_100"]
#             set_border_and_align(cell_info)
#             set_number_format(cell_info,"five_digit")


#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+4,column=columns_location["w1_monthly_view_value"])
#             cell_info.value = monthly_visit_kpi_w1
#             set_border_and_align(cell_info)
#             set_number_format(cell_info,"five_digit")


#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+4,column=columns_location["w1_kpi_achievement_value"])
#             # cell_info.value = float(phase_1["percent_100"])/float(monthly_view_kpi_w1_org)
#             set_border_and_align(cell_info)
#             set_number_format(cell_info,"percentage")


#             # ----------- Ending visits  ---------------------
#             # W1 above this ---------------------------------

#             # W2 starting below this ---------------------------------


#             # View starts here ------------------------------
#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary,column=columns_location["w2_heading_start"])
#             cell_info.value = "W2 KPI Summary"
#             set_border_and_red_bg(cell_info)
#             youtube_trueview_sheet. merge_cells(start_row=starting_row_Summary, start_column=columns_location["w2_heading_start"], end_row=starting_row_Summary, end_column=columns_location["w2_heading_end"])

#             try:

#                 cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+1,column=columns_location["w2_views_heading"])
#                 cell_info.value = "Views"
#                 set_border_and_black_bg(cell_info)
                
#                 cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+1,column=columns_location["w2_monthly_view_heading"])
#                 cell_info.value = "Monthly View KPI"
#                 set_border_and_black_bg(cell_info)
                
#                 cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+1,column=columns_location["w2_kpi_achievement_heading"])
#                 cell_info.value = "KPI Achievement"
#                 set_border_and_black_bg(cell_info)
                
#                 cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+2,column=columns_location["w2_views_value"])
#                 cell_info.value = phase_2["percent_100"]
#                 set_border_and_align(cell_info)
#                 set_number_format(cell_info,"five_digit")
                
#                 cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+2,column=columns_location["w2_monthly_view_value"])
#                 cell_info.value = monthly_view_kpi_w2_org
#                 set_border_and_align(cell_info)
#                 set_number_format(cell_info,"five_digit")
                
#                 cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+2,column=columns_location["w2_kpi_achievement_value"])
#                 cell_info.value = float(phase_2["percent_100"])/float(monthly_view_kpi_w2_org)
#                 set_border_and_align(cell_info)
#                 set_number_format(cell_info,"percentage")
            
#             except:
#                 pass
#             # --------------------------------------------

#             # Visit starts here

#             # From here visits starts ---------------------------------

            

#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+3,column=columns_location["w2_views_heading"])
#             cell_info.value = "Visits"
#             set_border_and_black_bg(cell_info)
            
#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+3,column=columns_location["w2_monthly_view_heading"])
#             cell_info.value = "Monthly Visit KPI"
#             set_border_and_black_bg(cell_info)

#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+3,column=columns_location["w2_kpi_achievement_heading"])
#             cell_info.value = "KPI Achievement"
#             set_border_and_black_bg(cell_info)

#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+4,column=columns_location["w2_views_value"])
#             # cell_info.value = phase_1["percent_100"]
#             set_border_and_align(cell_info)
#             set_number_format(cell_info,"five_digit")
            
#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+4,column=columns_location["w2_monthly_view_value"])
#             cell_info.value = monthly_visit_kpi_w2
#             set_number_format(cell_info,"five_digit")
#             set_border_and_align(cell_info)

#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+4,column=columns_location["w2_kpi_achievement_value"])
#             # cell_info.value = float(phase_1["percent_100"])/float(monthly_view_kpi_w1_org)
#             set_border_and_align(cell_info)
#             set_number_format(cell_info,"percentage")

#             # ----------- Ending visits  ---------------------

#             # -----------------------------------------------


#             # Total budgets - phase 1 and phase 2 

#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+6,column=columns_location["w1_heading_start"])
#             cell_info.value = "Youtube Trueview"
            

#             set_border_and_align_weekly(cell_info)
#             youtube_trueview_sheet. merge_cells(start_row=starting_row_Summary+6, start_column=columns_location["w1_heading_start"], end_row=starting_row_Summary+6, end_column=columns_location["w1_heading_end"])

#             # Left Headings 
#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+7,column=columns_location["w1_views_heading"])
#             cell_info.value = "Phase"
#             set_border_and_red_bg(cell_info)
            
            
#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+8,column=columns_location["w1_views_heading"])
#             cell_info.value = "Total Budget "
#             set_border_and_red_bg(cell_info)
#             # -----------------------------------

#             # Top Headings 
#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+7,column=columns_location["w1_monthly_view_heading"])
#             cell_info.value = "Phase 1"
#             set_border_and_align(cell_info)
            
#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+7,column=columns_location["w1_kpi_achievement_heading"])
#             cell_info.value = "Phase 2"
#             set_border_and_align(cell_info)
#             # -------------------------------------
           
#             # Values ------------------------------
#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+8,column=columns_location["w1_monthly_view_heading"])
#             cell_info.value = (phase_1_budget_org)
#             set_border_and_align(cell_info)
#             set_number_format(cell_info,"four_decimal_dollar")

#             cell_info = youtube_trueview_sheet.cell(row=starting_row_Summary+8,column=columns_location["w1_kpi_achievement_heading"])
#             cell_info.value = (phase_2_budget_org)
#             set_border_and_align(cell_info)
#             set_number_format(cell_info,"four_decimal_dollar")
#             # ---------------------------------------


#             # Table with day wise and phases 

#             row_num+=1
#             row_num+=1

#             for key,value in columns_label.items():
#                 cell_location = youtube_trueview_sheet.cell(row=row_num, column=value)
#                 cell_location.value =  key     # This is the title of the table with title : Date
#                 cell_location.border = thin_border
#                 cell_location.alignment = TEXT_ALIGNMENT
#                 cell_location.fill = PatternFill(start_color="DAEEF3",end_color="DAEEF3", fill_type = "solid")

#             # Printing column names in excel sheet     
                
#             weekly_counter = 0              # In the start weekly counter will be 0 
#             weekly_number = 0               # Weekly number will be 0 
#             row_num = row_num + 1                     # Data storing location by row
#             previous_budget = 0  
#             total_weekly_kpi = 0 
#             previous_budget_by_week = 0
#             total_phase_1_daily_kpi= 0

#             weekly_data = []
#             # First thing is to run a loop to store the total data  
#             for i in range(2,len(df)+2):            # Loop to start adding in the file 
#                     cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Date"])
#                     cell_info.value = df.index[i-2]    # This the date which is storing in the file  
#                     set_border_and_align(cell_info)

#                     cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Impressions"])
#                     cell_info.value = df["Impressions"][i-2]           # This is the impressions storing in the file 
#                     set_number_format(cell_info,"comma_format")
#                     set_border_and_align(cell_info)

#                     cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Clicks"])
#                     cell_info.value = df["Clicks"][i-2]                # This is the clicks storing in the file 
#                     set_border_and_align(cell_info)
#                     set_number_format(cell_info,"comma_format")

#                     try:
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Three Views"])
#                         cell_info.value = df["Three Views"][i-2]           # This is the Three view storing in the file  
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"comma_format")
#                     except:
#                         pass

#                     try:

#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Trueview"])
#                         cell_info.value = df["Trueviews"][i-2]           # This is the true view storing in the file 
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"comma_format")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["25_percent"])
#                         cell_info.value = df["25% Views"][i-2]          # This is the 25% view 
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"comma_format")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["50_percent"])
#                         cell_info.value = df["50% Views"][i-2]                # This is the 50% view
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"comma_format")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["75_percent"])
#                         cell_info.value = df["75% Views"][i-2]          # This is the 75% view
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"comma_format")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["100_percent"])
#                         cell_info.value = df["100% Views"][i-2]                # This is the 100% view
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"comma_format")

#                     except:
#                         pass   
#                     cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["spendings"])
                    
#                     cell_info.value = df["Spendings"][i-2]                # This is the spending 
#                               # This is the spending 
                        
#                     set_border_and_align(cell_info)
#                     set_number_format(cell_info,"two_decimal_dollar")
                                        
#                     cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["budget_spent"])
                    
#                     cell_info.value = (previous_budget+ df["Spendings"][i-2]) / phase_1_budget                 # This is the spending 
   
#                     set_border_and_align(cell_info)
#                     set_number_format(cell_info,"percentage")
                    
           
#                     previous_budget = previous_budget+ df["Spendings"][i-2]
               
#                     total_weekly_kpi = total_weekly_kpi + (monthly_view_kpi_w1 / 30)
#                     total_phase_1_daily_kpi = total_phase_1_daily_kpi + monthly_view_kpi_w1 / 30 
                    
                    
#                     cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["ctr"])
#                     cell_info.value = df["Clicks"][i-2] / df["Impressions"][i-2]                # This is the CTR 
#                     set_border_and_align(cell_info)
#                     set_number_format(cell_info,"two_decimal_percentage")
#                     try:
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["cpv_3_views"])
#                         cell_info.value = df["Spendings"][i-2]  / df["Three Views"][i-2] 
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"four_decimal_dollar")
#                     except:
#                         pass


#                     try:
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["daily_view_kpi"])
#                         cell_info.value = monthly_view_kpi_w1 / 30               # This is the Daily View KPI 
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"five_digit")
                        
                        

#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["cpv_trueview"])
#                         cell_info.value = df["100% Views"][i-2]  / df["Trueviews"][i-2] 
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"four_decimal_dollar")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["cpv_complete"])
#                         cell_info.value = df["Spendings"][i-2]  / df["100% Views"][i-2] 
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"four_decimal_dollar")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["daily_kpi_achievement"])
#                         cell_info.value = df["100% Views"][i-2]  / (monthly_view_kpi_w1 / 30) 
#                         # Add previous sum in the new and divide by actual budget 
#                         set_border_and_align(cell_info)
#                         set_number_format(cell_info,"percentage")
#                     except:
#                         pass

#                     weekly_counter +=1              # counting the values and checking if the records are exceeding and getting more than week then print week 
#                     if weekly_counter >=7 or i == len(df)+1:          # if the weekly counter is greater than or equal to 7 then print week 
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Date"])
#                         cell_info.value = "Weekly Total"             # Printing weekly Total
#                         set_border_and_align_weekly(cell_info)
#                         this_week_start_date = weekly_list[weekly_number]['date_start']
#                         this_week_end_date = weekly_list[weekly_number]['date_end']
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Impressions"])
#                         cell_info.value = weekly_list[weekly_number]['impressions']      # Printing impressions for week
#                         # annual_total["impressions"]+= weekly_list[weekly_number]['impressions']
#                         set_border_and_align_weekly(cell_info)
#                         set_number_format(cell_info,"comma_format")
#                         this_week_impressions = cell_info.value

                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Clicks"])
#                         cell_info.value = weekly_list[weekly_number]['clicks']  # Printing clicks for week
#                         # annual_total["clicks"]+= weekly_list[weekly_number]['clicks']
#                         set_border_and_align_weekly(cell_info)
#                         set_number_format(cell_info,"comma_format")
#                         this_week_clicks = cell_info.value

#                         try:
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Three Views"])
#                             cell_info.value = weekly_list[weekly_number]['three_views'] # Printing Three views for week
#                             set_border_and_align_weekly(cell_info)
#                             set_number_format(cell_info,"comma_format")
#                             this_week_three_view = cell_info.value


#                         except:
#                             this_week_three_view = "0"
#                             pass

#                         try:
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Trueview"])
#                             cell_info.value = weekly_list[weekly_number]['trueview'] # Printing trueview for week
#                             # annual_total["trueview"]+= weekly_list[weekly_number]['trueview']
#                             set_border_and_align_weekly(cell_info)
#                             set_number_format(cell_info,"comma_format")
#                             this_week_trueview = cell_info.value
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["25_percent"])
#                             cell_info.value = weekly_list[weekly_number]['percent_25']  # Printing 25% views for week
#                             # annual_total["25_percentage_view"]+= weekly_list[weekly_number]['percent_25']
#                             set_border_and_align_weekly(cell_info)
#                             set_number_format(cell_info,"comma_format")
#                             this_week_25_percent = cell_info.value
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["50_percent"])
#                             cell_info.value = weekly_list[weekly_number]['percent_50'] # Printing 50% views for week
#                             # annual_total["50_percentage_view"]+= weekly_list[weekly_number]['percent_50']
#                             set_border_and_align_weekly(cell_info)
#                             set_number_format(cell_info,"comma_format")
#                             this_week_50_percent = cell_info.value
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["75_percent"])
#                             cell_info.value = weekly_list[weekly_number]['percent_75']    # Printing 75% for week
#                             # annual_total["75_percentage_view"]+= weekly_list[weekly_number]['percent_75']
#                             set_border_and_align_weekly(cell_info)
#                             set_number_format(cell_info,"comma_format")
#                             this_week_75_percent = cell_info.value
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["100_percent"])
#                             cell_info.value = weekly_list[weekly_number]['percent_100']       # Printing 100% for week
#                             # annual_total["100_percentage_view"]+= weekly_list[weekly_number]['percent_100']
#                             set_border_and_align_weekly(cell_info)
#                             set_number_format(cell_info,"comma_format")
#                             this_week_100_percent = cell_info.value


#                         except:
#                             pass
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["spendings"])
#                         cell_info.value = weekly_list[weekly_number]['spending']      # Printing spendings for week
#                         # annual_total["spending"]+= weekly_list[weekly_number]['spending']
#                         set_border_and_align_weekly(cell_info)
#                         set_number_format(cell_info,"two_decimal_dollar")
#                         This_week_spendings = cell_info.value

                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["budget_spent"])
#                         cell_info.value = (previous_budget_by_week + weekly_list[weekly_number]['spending'])  / phase_1_budget                # This is the Budget Spent 
#                         set_border_and_align_weekly(cell_info)
#                         set_number_format(cell_info,"percentage")
#                         This_week_budget_spent = cell_info.value
                        
                        
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["ctr"])
#                         cell_info.value = weekly_list[weekly_number]['clicks'] / weekly_list[weekly_number]['impressions']                # This is the CTR 
#                         set_border_and_align_weekly(cell_info)
#                         set_number_format(cell_info,"two_decimal_percentage")
#                         This_week_ctr = cell_info.value

                        
#                         try:
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["cpv_3_views"])
#                             cell_info.value = weekly_list[weekly_number]['spending']   / weekly_list[weekly_number]['three_views'] 
#                             set_border_and_align_weekly(cell_info)
#                             set_number_format(cell_info,"four_decimal_dollar")
#                             This_week_cpv_3_views = cell_info.value


#                         except:
#                             This_week_cpv_3_views = "-"
                            
#                             pass

#                         try:
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["daily_view_kpi"])
#                             cell_info.value = total_weekly_kpi                   # This is the Daily View KPI 
#                             # annual_total["weekly_view_kpi"] = annual_total["weekly_view_kpi"] + total_weekly_kpi
#                             set_border_and_align_weekly(cell_info)
#                             set_number_format(cell_info,"five_digit")
#                             This_week_daily_view_kpi = cell_info.value

                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["cpv_trueview"])
#                             cell_info.value = weekly_list[weekly_number]['percent_100']   / weekly_list[weekly_number]['trueview'] 
#                             set_border_and_align_weekly(cell_info)
#                             set_number_format(cell_info,"four_decimal_dollar")
#                             This_week_cpv_trueview = cell_info.value

                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["cpv_complete"])
#                             cell_info.value =   weekly_list[weekly_number]['spending']  / weekly_list[weekly_number]['percent_100'] 
#                             set_border_and_align_weekly(cell_info)
#                             set_number_format(cell_info,"four_decimal_dollar")
#                             This_week_cpv_complete = cell_info.value


#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["daily_kpi_achievement"])
#                             cell_info.value =   weekly_list[weekly_number]['percent_100']  / total_weekly_kpi
#                             set_border_and_align_weekly(cell_info)
#                             set_number_format(cell_info,"percentage")
#                             This_week_Daily_KPI_Achievement = cell_info.value
#                             # print(weekly_list[weekly_number]['percent_100']," : ",total_weekly_kpi)
#                         except:
#                             This_week_Daily_KPI_Achievement = "-"
#                             pass
                        
#                         this_week = {
#                             "start_date": this_week_start_date,
#                             "end_date": this_week_end_date,
#                             "impressions": this_week_impressions,
#                             "clicks": this_week_clicks,
#                             "three_view": this_week_three_view,
#                             "trueview":this_week_trueview,
#                             "percent_25_views": this_week_25_percent,
#                             "percent_50_views": this_week_50_percent,
#                             "percent_75_views": this_week_75_percent,
#                             "percent_100_views": this_week_100_percent,
#                             "spendings": This_week_spendings,
#                             "budget_spent": This_week_budget_spent,
#                             "ctr": This_week_ctr,
#                             "cpv_three_views": This_week_cpv_3_views,
#                             "daily_view_kpi": This_week_daily_view_kpi,
#                             "cpv_trueview": This_week_cpv_trueview,
#                             "cpv_complete": This_week_cpv_complete,
#                             "daily_kpi_achievement": This_week_Daily_KPI_Achievement

#                         }
#                         weekly_data.append(this_week)
#                         previous_budget_by_week=previous_budget_by_week + weekly_list[weekly_number]['spending']

                        
                        
                        
#                         total_weekly_kpi=  0
#                         weekly_number +=1           # incrementing week 
#                         row_num+=1                  # Incrementing the row so we can store afterwards record 
#                         weekly_counter = 0          # making the counter back to 0 
                    
#                     # Now we are ready to deal with phase 1 and phase 2 
#                     # So if the wee
#                     # print(phase_1_deadline, df.index[i-2])
#                     if phase_1_deadline == df.index[i-2]:
#                         # print("----------------------------------------phase_1 ----------------------------------------------")
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Date"])
#                         cell_info.value = "Phase 1 Total"             # Printing weekly Total
#                         set_border_and_align_phase(cell_info)
#                         set_number_format(cell_info,"comma_format")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Impressions"])
#                         cell_info.value = phase_1['impressions']      # Printing impressions for week
#                         set_border_and_align_phase(cell_info)
#                         set_number_format(cell_info,"comma_format")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Clicks"])
#                         cell_info.value = phase_1['clicks']  # Printing clicks for week
#                         set_border_and_align_phase(cell_info)
#                         set_number_format(cell_info,"comma_format")
                        
#                         try:
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Three Views"])
#                             cell_info.value = phase_1['three_views'] # Printing three views for week
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"comma_format")

#                         except:
#                             pass

#                         try:
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Trueview"])
#                             cell_info.value = phase_1['trueview'] # Printing trueview for week
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"comma_format")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["25_percent"])
#                             cell_info.value = phase_1['percent_25']  # Printing 25% views for week
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"comma_format")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["50_percent"])
#                             cell_info.value = phase_1['percent_50'] # Printing 50% views for week
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"comma_format")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["75_percent"])
#                             cell_info.value = phase_1['percent_75']    # Printing 75% for week
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"comma_format")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["100_percent"])
#                             cell_info.value = phase_1['percent_100']       # Printing 100% for week
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"comma_format")

#                         except:
#                             pass
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["spendings"])
#                         cell_info.value = phase_1['spending']      # Printing spendings for week
#                         set_border_and_align_phase(cell_info)
#                         set_number_format(cell_info,"two_decimal_dollar")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["budget_spent"])
#                         cell_info.value = previous_budget /  phase_1_budget            # This is the spending 
#                         set_border_and_align_phase(cell_info)
#                         set_number_format(cell_info,"percentage")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["ctr"])
#                         cell_info.value = phase_1['clicks'] / phase_1['impressions']       # Printing spendings for week
#                         set_border_and_align_phase(cell_info)
#                         set_number_format(cell_info,"two_decimal_percentage")
                        
#                         try:
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["cpv_3_views"])
#                             cell_info.value = phase_1['spending']   / phase_1['three_views'] 
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"four_decimal_dollar")
#                         except:
#                             pass

#                         try:
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["daily_view_kpi"])
#                             cell_info.value = total_phase_1_daily_kpi                   # This is the spending 
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"five_digit")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["cpv_trueview"])
#                             cell_info.value = phase_1['percent_100']   / phase_1['trueview'] 
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"four_decimal_dollar")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["cpv_complete"])
#                             cell_info.value = phase_1['spending']     / phase_1['percent_100'] 
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"four_decimal_dollar")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["daily_kpi_achievement"])
#                             cell_info.value = phase_1['trueview']      / total_phase_1_daily_kpi
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"percentage")
#                         except:
#                             pass

#                         phase_1_budget = phase_2_budget
#                         total_phase_1_daily_kpi = 0 
#                         monthly_view_kpi_w1 = monthly_view_kpi_w2
#                         row_num+=1  
#                         previous_budget =0  
#                         previous_budget_by_week =0  
                    
#                     if phase_2_deadline == df.index[i-2]:
#                         # print("phase_1")
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Date"])
#                         cell_info.value = "Phase 2 Total"             # Printing weekly Total
#                         set_border_and_align_phase(cell_info)
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Impressions"])
#                         cell_info.value = phase_2['impressions']      # Printing impressions for week
#                         set_border_and_align_phase(cell_info)
#                         set_number_format(cell_info,"comma_format")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Clicks"])
#                         cell_info.value = phase_2['clicks']  # Printing clicks for week
#                         set_border_and_align_phase(cell_info)
#                         set_number_format(cell_info,"comma_format")
                        
                        
#                         try:
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Three Views"])
#                             cell_info.value = phase_2['three_views'] # Printing Three views for week
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"comma_format")
#                         except:
#                             pass


#                         try:



#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["Trueview"])
#                             cell_info.value = phase_2['trueview'] # Printing trueview for week
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"comma_format")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["25_percent"])
#                             cell_info.value = phase_2['percent_25']  # Printing 25% views for week
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"comma_format")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["50_percent"])
#                             cell_info.value = phase_2['percent_50'] # Printing 50% views for week
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"comma_format")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["75_percent"])
#                             cell_info.value = phase_2['percent_75']    # Printing 75% for week
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"comma_format")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["100_percent"])
#                             cell_info.value = phase_2['percent_100']       # Printing 100% for week
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"comma_format")

#                         except:
#                             pass

#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["spendings"])
#                         cell_info.value = phase_2['spending']      # Printing spendings for week
#                         set_border_and_align_phase(cell_info)
#                         set_number_format(cell_info,"two_decimal_dollar")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["budget_spent"])
#                         cell_info.value = previous_budget /  phase_2_budget      # Printing spendings for week
#                         set_border_and_align_phase(cell_info)
#                         set_number_format(cell_info,"percentage")
                        
#                         cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["ctr"])
#                         cell_info.value = phase_2['clicks'] / phase_2['impressions']       # Printing spendings for week
#                         set_border_and_align_phase(cell_info)
#                         set_number_format(cell_info,"two_decimal_percentage")
                        
#                         try:
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["cpv_3_views"])
#                             cell_info.value = phase_2['spending'] / phase_2['three_views']                # This is the spending 
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"four_decimal_dollar")
#                         except:
#                             pass

#                         try:

#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["daily_view_kpi"])
#                             cell_info.value = total_phase_1_daily_kpi                   # This is the spending 
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"five_digit")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["cpv_trueview"])
#                             cell_info.value = phase_2['percent_100'] / phase_2['trueview']                # This is the spending 
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"four_decimal_dollar")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["cpv_complete"])
#                             cell_info.value = phase_2['spending']     / phase_2['percent_100'] 
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"four_decimal_dollar")
                            
#                             cell_info = youtube_trueview_sheet.cell(row=row_num+1, column=columns_location["daily_kpi_achievement"])
#                             cell_info.value = phase_2['trueview'] / total_phase_1_daily_kpi              # This is the spending 
#                             set_border_and_align_phase(cell_info)
#                             set_number_format(cell_info,"percentage")
#                         except:
#                             pass
                    
#                         row_num+=1   
#                     row_num+=1

            
#             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Date"])
#             cell_info.value = "Total"             # Printing weekly Total
#             set_border_and_red_bg(cell_info)
            
#             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Impressions"])
#             cell_info.value = annual_total['impressions']      # Printing impressions for week
#             set_border_and_red_bg(cell_info)
#             set_number_format(cell_info,"comma_format")
            
#             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Clicks"])
#             cell_info.value = annual_total['clicks']  # Printing clicks for week
#             set_border_and_red_bg(cell_info)
#             set_number_format(cell_info,"comma_format")
            
#             try:

#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Three Views"])
#                 cell_info.value = annual_total['three_views'] # Printing three views for week
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"comma_format")

#             except:
#                 pass
            
#             try:

#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["Trueview"])
#                 cell_info.value = annual_total['trueview'] # Printing trueview for week
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"comma_format")
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["25_percent"])
#                 cell_info.value = annual_total['25_percentage_view']  # Printing 25% views for week
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"comma_format")
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["50_percent"])
#                 cell_info.value = annual_total['50_percentage_view'] # Printing 50% views for week
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"comma_format")
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["75_percent"])
#                 cell_info.value = annual_total['75_percentage_view']    # Printing 75% for week
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"comma_format")
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["100_percent"])
#                 cell_info.value = annual_total['100_percentage_view']       # Printing 100% for week
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"comma_format")
#             except:
#                 pass       
#             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["spendings"])
#             cell_info.value = annual_total['spending']      # Printing spendings for week
#             set_border_and_red_bg(cell_info)
#             set_number_format(cell_info,"two_decimal_dollar")
            
#             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["budget_spent"])
#             cell_info.value = annual_total["budget_spent"]      # Printing spendings for week
#             set_border_and_red_bg(cell_info)
#             set_number_format(cell_info,"percentage")
            
#             cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["ctr"])
#             cell_info.value = annual_total["ctr"]       # Printing spendings for week
#             set_border_and_red_bg(cell_info)
#             set_number_format(cell_info,"two_decimal_percentage")
#             try:
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["cpv_3_views"])
#                 cell_info.value = annual_total["cpv_3_views"]               # This is the spending 
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"four_decimal_dollar")
                
#             except:
#                 pass

#             try:
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["daily_view_kpi"])
#                 cell_info.value = annual_total["weekly_view_kpi"]                  # This is the spending 
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"five_digit")
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["cpv_trueview"])
#                 cell_info.value = annual_total["cpv_trueview"]               # This is the spending 
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"four_decimal_dollar")
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["cpv_complete"])
#                 cell_info.value = annual_total["cpv_complete"] 
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"four_decimal_dollar")
                
#                 cell_info = youtube_trueview_sheet.cell(row=row_num, column=columns_location["daily_kpi_achievement"])
#                 cell_info.value = annual_total["weekly_kpi_achievement"]           # This is the spending 
#                 set_border_and_red_bg(cell_info)
#                 set_number_format(cell_info,"percentage")

#             except:
#                 pass
            

                    

#             # Store the weekly data 
#             # Store the phases 
#             # Store in the sheet 
#             # Return the file 
#             # default_template.save(str(filename)+".xlsx")

#             # Storing file in the database   --------------------
#             serialized_workbook = pickle.dumps(default_template)
#             uploading_file = media_files(filename,serialized_workbook,current_user.username)
#             db.session.add(uploading_file)
#             db.session.commit()
#             db.session.refresh(uploading_file)
#             file_id = uploading_file.id
#             # print(weekly_data)
#             print("----ANNUAL RECORD----")
#             print(annual_total)
#             print("-- -- -- -- -- -- --")

#             create_annual = annual_record(annual_total,file_id)
#             db.session.add(create_annual)
#             db.session.commit()

#             for i in range(0,(len(weekly_data))):
#                 # this_week = {
#                 #             "start_date": this_week_start_date,
#                 #             "end_date": this_week_end_date,
#                 #             "impressions": this_week_impressions,
#                 #             "clicks": this_week_clicks,
#                 #             "three_view": this_week_three_view,
#                 #             "trueview":this_week_trueview,
#                 #             "percent_25_views": this_week_25_percent,
#                 #             "percent_50_views": this_week_50_percent,
#                 #             "percent_75_views": this_week_75_percent,
#                 #             "percent_100_views": this_week_100_percent,
#                 #             "spendings": This_week_spendings,
#                 #             "budget_spent": This_week_budget_spent,
#                 #             "ctr": This_week_ctr,
#                 #             "cpv_three_views": This_week_cpv_3_views,
#                 #             "daily_view_kpi": This_week_daily_view_kpi,
#                 #             "cpv_trueview": This_week_cpv_trueview,
#                 #             "cpv_complete": This_week_cpv_complete,
#                 #             "daily_kpi_achievement": This_week_Daily_KPI_Achievement

#                 #         }
#                 week_ = weekly_record(
#                     weekly_data[i]['start_date'],
#                     weekly_data[i]['end_date'],
#                     weekly_data[i]['impressions'],
#                     weekly_data[i]['clicks'],
#                     weekly_data[i]['trueview'],
#                     weekly_data[i]['spendings'],
#                     weekly_data[i]['percent_25_views'],
#                     weekly_data[i]['percent_50_views'],
#                     weekly_data[i]['percent_75_views'],
#                     weekly_data[i]['percent_100_views'],
#                     weekly_data[i]['budget_spent'],
#                     weekly_data[i]['ctr'],
#                     weekly_data[i]['daily_view_kpi'],
#                     weekly_data[i]['cpv_trueview'],
#                     weekly_data[i]['cpv_complete'],
#                     weekly_data[i]['daily_kpi_achievement'],
#                     weekly_data[i]['three_view'],
#                     weekly_data[i]['cpv_three_views'],
#                     file_id,
#                 )
#                 db.session.add(week_)
#                 db.session.commit()








            
#             db.session.close()

#             # Storing completed ---------------------
#             return jsonify({
#                 "Message" : "Report Generated Successfully",
#                 "status" : 200
#             },200)
#             # return send_file(str(filename)+".xlsx")
#             # return render_template("youtube.html")
#         # else:
#         #     return render_template("youtube.html")

# def Weekly_Total(weekly_,MVKPI,totalBudget,type):

#     # print("Correct")
#     weekly_total_ = weekly_.copy()
#     # print("working-till 1st correct")

#     weekly_total_.drop(weekly_total_.index, inplace=True)
#     # print("working-till 2nd correct")
#     # print(gdn_df_daily)
#     for i in range(0, 1):
#         row = weekly_.iloc[i:len(weekly_)].select_dtypes(include=['int64','double']).sum()
        
#         row['CTR'] = row['Clicks'] / row['Impressions']
#         # print("CTR")
#         row['Budget'] = round(row['Budget'],2)
#         row['type'] = type
#         row['SPENT_BUDGET'] = round((row['Budget']/int(totalBudget))*100)
#         # if row['3 Views'] == None or row['3 Views'] == 0:
#         #     row['3 Views'] = row['TrueView: Views']

        
#         # row['DAILY_KPI_ACHIEVEMENT'] = round(int(MVKPI) / row['3 Views'])  if row['3 Views'] !=0 and row['3 Views'] != None else  round(int(MVKPI) / row['View 100%'])
#         row['CTR'] = round(row['CTR']*100,2)
#         # if row['3 Views'] == None or row['3 Views'] == 0:
#         #     pass
#         # else:
#         try:
#             row['MONTHLY_KPI_ACHIEVEMENT'] =  row['3 Views'] / int(MVKPI)
#             # print ("Values: ", row['3 Views'])
#         except:
#             row['MONTHLY_KPI_ACHIEVEMENT'] =  row['View 100%'] / int(MVKPI)
#             # print("Values", row['TrueView: Views'])
#         # row['MONTHLY_KPI_ACHIEVEMENT'] =  row['3 Views'] / int(MVKPI)
#         # print("CTR_AGAINS")
#         # row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
#         # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
#         weekly_total_ = weekly_total_.append(row, ignore_index=True)
#         # print("Append row")
#     #
#     # print("working-till loop correct")
#     # del weekly_total_['DAILY_KPI_ACHIEVEMENT']
#     # weekly_total_ = weekly_total_.assign(WEEKLY_KPI_ACHIEVEMENT = (round(int(MVKPI) / weekly_total_['3 Views'])))
#     # weekly_total_ = weekly_total_.assign(SPENT_BUDGET_TOTAL = round((weekly_total_['Budget']/totalBudget)*100,2))
#     weekly_total_ = weekly_total_.assign(MVKPI = int(MVKPI))
#     # weekly_total_ = weekly_total_.assign(MONTHLY_KPI_ACHIEVEMENT = )
#     # print("working-till assign correct")
    
#     # if i == 0:
#     #     weekly_total_.to_sql("summary",con=db.engine,if_exists="replace",index=False)
#     # else:
#     #     weekly_total_.to_sql("summary",con=db.engine,if_exists="append",index=False)
#     # print("Success")
#     return weekly_total_


# def SettingCTR_TRUEVIEW(df,identifier,budget,monthly_view_kpi):
#             # try:
#                 df['Budget'] = (round(df['Budget'],2)).replace([np.inf, -np.inf,np.nan], 0)
#                 df = df.assign(DAILY_VIEW_KPI = (round(float(monthly_view_kpi) / 30)))
#                 df = df.assign(DAILY_KPI_ACHIEVEMENT = (round((df['View 100%'] / df['DAILY_VIEW_KPI'])*100)).replace([np.inf, -np.inf], 0))
#                 df = df.assign(SPENT_BUDGET = (round((df['Budget'].cumsum()/int(budget))*100)).replace([np.inf, -np.inf], 0))
#                 df = df.assign(CPV_COMPLETE = (round(df["Budget"]/df["View 100%"],4)).replace([np.inf, -np.inf], 0))
#                 df = df.assign(CTR = (round(((df["Clicks"]/df["Impressions"])*100),2).replace([np.inf, -np.inf], 0)))
#                 if identifier == "fb_video" or identifier == "sc_video" or identifier == "ig_video" or identifier == "twitter_video" or identifier =="fb_image" or identifier == "sc_image" or identifier == "ig_image" or identifier == "twitter_image":
#                     print("coming here",identifier)
#                     df = df.assign(CPV_TRUEVIEW = (round(df["Budget"]/df["ThruPlays"],4)).replace([np.inf, -np.inf], 0))
#                 else:
#                     df = df.assign(CPV_TRUEVIEW = (round(df["Budget"]/df["TrueView: Views"],4)).replace([np.inf, -np.inf], 0))
                
                    
                    
#             # except:
#                 # print("ERROR HERE AS WELL")
#                 return df
# def CalculatingWeekly(df_daily,type):
#     date_rng = pd.date_range(start=df_daily.index.min(), end=df_daily.index.max(), freq='D')
#     # print("Fine")
#     # Fetching the start and end date so we can loop through it for weeks 
#     df_weekly = df_daily.copy()
#     # Making a copy of daily table so we can use the same columns for weekly as well
#     # print("Still fine")
#     df_weekly.drop(df_weekly.index, inplace=True)
#     # We are dropping all the data so we can add new data inside weekly 
#     # print("still still fine")

#     # print(gdn_df_daily)
#     # Starting the loop for weeks 
#     for i in range(0, len(date_rng), 7):
#         try:
#             row = df_daily.iloc[i:i+7].select_dtypes(include=['int64','double']).sum()
#         except:
#             print("not fine 1")
#         # Doing sum for the row from 1 - 7 days and storing in row 
#         try:
#             row['Budget'] = round(row['Budget'],2)
#         # Changing the Format of the budget 
#         except:
#             print("not fine 2")
#         try:
#             row['CTR']=round(((row['Clicks']/row['Impressions'])*100),2)
#         except:
#             print("not fine 3")
#         # Changing the CTR for the week 
#         # row['CTR'] = round(row['CTR'],2)
#         # Changing the format for CTR

#         try:
#             row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
#         except:
#             print("not fine 4")
        
#         try:
#             try:
#                 if row["ThruPlays"] == 0:
#                     row['CPV_TRUEVIEW'] = 0
#                 else:
#                     row['CPV_TRUEVIEW'] = round(row["Budget"]/row["ThruPlays"],4)

#             except:
#                 if row["TrueView: Views"] == 0:
#                     row['CPV_TRUEVIEW'] = 0
#                 else:
#                     row['CPV_TRUEVIEW'] = round(row["Budget"]/row["TrueView: Views"],4)
#         except:
#             print("not fine 5")
#         # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
#         if row["View 100%"] == 0:
#             row['CPV_COMPLETE'] = 0
#         else:
#             row['CPV_COMPLETE'] = round(row["Budget"]/row["View 100%"],4)

#         if i+6 > len(date_rng):
          
#             row["SPENT_BUDGET"] = df_daily["SPENT_BUDGET"][len(date_rng)-1]
#         else : 
#             # row['Budget'] = round(row['Budget'],2)
#             row["SPENT_BUDGET"] = df_daily["SPENT_BUDGET"][i+6]
#             # row['CTR'] = round(row['CTR'],2)
#             # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
#             # row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
#             row["DAILY_KPI_ACHIEVEMENT"] = round(row["View 100%"]/ row["DAILY_VIEW_KPI"]*100)
#             # if type == "fb_video" or type == "fb_image" or type=="ig_image" or type=="ig_video":
#             # else:
#             #     row["DAILY_KPI_ACHIEVEMENT"] = round(row["View 100%"]/ row["DAILY_VIEW_KPI"]*100)
#         df_weekly = df_weekly.append(row, ignore_index=True)
#     # print("absolutely fine")
#     # del df_weekly['type'] 
#     return df_weekly 
# def DataFrameToSQL(df,identifier,identifier_name):
#     if identifier_name == "fb_video" or identifier_name == "fb_image" or identifier_name == "sc_video" or identifier_name == "sc_image" or identifier_name == "twitter_video" or identifier_name == "twitter_image" or identifier_name == "ig_video"or identifier_name == "ig_image":
#         # print(identifier_name)
#         # print(df["Campaign name"].str.contains(identifier))
#         df = df[df["Campaign name"].str.contains(identifier)]
#     else:
#         print(identifier_name)
#         try:
#             df =df[df["Insertion Order"].str.contains(identifier)]  
#         except:
#              pass

#     df = df.assign(type = identifier_name)
#     return df
#         # df.insert(1, 'type', identifier_name, True)
#         # try:
#         #     # df.to_sql(filename,con=db.engine,if_exists="fail",index=True,index_label="id")
#         #     return df,True
#         # except:
#         #     db.engine.execute(f"DROP table IF EXISTS `{filename}`")
#         #     pass
#         #     # If we want to delete a table we will use above line 
#         # return df,False



# def StylingSheets(writer,sheet_name):
#     left_border = Side(style='thin', color='000000')
#     right_border = Side(style='thin', color='000000')
#     top_border = Side(style='thin', color='000000')
#     bottom_border = Side(style='thin', color='000000')

#     # Create the border
#     border = Border(left=left_border, right=right_border, top=top_border, bottom=bottom_border)
#     workbook = writer.book
#     worksheet = writer.sheets[sheet_name]
#     row_heading = 0
#     for row in worksheet.iter_rows():
#         empty_count = 0 
#         if row_heading == 0:
#             row_heading+=1
         
#             for cell in row:
#                 if not cell.value:
#                     # print("this", empty_count)
#                     empty_count +=1

#             if empty_count < len(row)-1:
             
#                 for cell in row:
#                     cell.font = Font(size=10, bold=True)
#                     cell.font = cell.font.copy(name='맑은 고딕')
#                     cell.alignment = Alignment(horizontal='center')
#                     cell.alignment = Alignment(vertical='center')
#                     cell.alignment = cell.alignment.copy(wrap_text=True)
#                     # cell.width = None
#             else:
#                     row_heading = 0
#         else:
#             for cell in row:
#                 if not cell.value:
#                     # print("this", empty_count)
#                     empty_count +=1

#             if empty_count < len(row)-1:
          
#                 for cell in row:
#                 # cell.fill = PatternFill("solid", fgColor="DDDDDD")
#                 # cell.alignment = Alignment(horizontal='center')
#                     cell.font = Font(size=10, bold=False)
#                     cell.font = cell.font.copy(name='맑은 고딕')
#                     cell.alignment = Alignment(horizontal='center')
#                     cell.alignment = Alignment(vertical='center')
#                     cell.border = border
#                     cell.alignment = cell.alignment.copy(wrap_text=True)
#                     # cell.width = None
#             else:
#                 row_heading = 0
#             # cell.alignment = Alignment(horizontal='center')
#             # cell.alignment = Alignment(vertical='center')

#     # for column_cells in worksheet.columns:
#     #     print("running")
#     #     worksheet.column_dimensions[column_cells[0].column_letter].width = None

#     return writer


# # Route: Raw Data Storage 
# # Description: In this route we will store raw data in database
# @app.route("/raw_data_storage",methods=["POST"])
# @cross_origin()
# @token_required

# def raw_data_storage(current_user):

#     filename = request.form["filename"]
#     # File name to be 
#     fb_video = request.form['fb_video']
#     fb_video_budget = request.form["fb_video_budget"]
#     fb_image = request.form['fb_image']
#     fb_image_budget = request.form["fb_image_budget"]


#     ig_video = request.form['ig_video']
#     ig_video_budget = request.form["ig_video_budget"]
#     ig_image = request.form['ig_image']
#     ig_image_budget = request.form["ig_image_budget"]
    
#     gdn = request.form['gdn']
#     gdn_budget = request.form["gdn_budget"]
    
#     trueview = request.form['trueview']
#     trueview_budget = request.form["trueview_budget"]
#     mvkpi_trueview = request.form['mvkpi_trueview']
#     mvkpi_gdn = request.form['mvkpi_gdn']
#     mvkpi_fb_video = request.form['mvkpi_fb_video']
#     mvkpi_fb_image = request.form['mvkpi_fb_image']
#     mvkpi_ig_image = request.form['mvkpi_ig_image']
#     mvkpi_ig_video = request.form['mvkpi_ig_video']

#     raw_data_file_fb_ig_snapchat = request.files["file-fb-ig-snapchat"]
#     raw_data_file_trueview_gdn = request.files["file-trueview-gdn"]


#     # Taking the files and the data as Input and saving in variable 


#     # We are getting the files 


#     df_fb_ig_snapchat = pd.read_excel(raw_data_file_fb_ig_snapchat)
#     df_trueview_gdn = pd.read_excel(raw_data_file_trueview_gdn)
#     print(df_trueview_gdn)
#     # We are reading the excel files 


#     # Getting dataframe for each date. 
#     # Getting status that either the data is stored or not 
#     # If data is stored then get the sum of the values 
#     # And store the calculated data in the database table -- ctr, and other calcualtion of the table 
#     # 
#     filename_youtube_gdn = filename + "-youtube_gdn" 
#     filename_fb_ig_snapchat = filename + "-fb_ig_snapchat" 
#     # Stored file names of two different categories

#     # We are adding Dataframes to SQL Table 
#     fb_video_df = DataFrameToSQL(df_fb_ig_snapchat,fb_video,"fb_video")
#     fb_image_df = DataFrameToSQL(df_fb_ig_snapchat,fb_image,"fb_image")
#     ig_video_df = DataFrameToSQL(df_fb_ig_snapchat,ig_video,"ig_video")
#     ig_image_df = DataFrameToSQL(df_fb_ig_snapchat,ig_image,"ig_image")
#     # -------------------------------------------------------------
#     gdn_df = DataFrameToSQL(df_trueview_gdn,gdn,"gdn")
#     trueview_df = DataFrameToSQL(df_trueview_gdn,trueview,"trueview")
#     # Assigning Types to raw data and returning the dataframe to be stored in SQL Table

#     # workbook = Workbook()
#     # trueview_df.to_excel(workbook,"trueview_test",index=False)
#     # try:
#     # Raw data stored in the raw data table 
#     fb_video_df.to_sql(filename_fb_ig_snapchat,con=db.engine,if_exists="replace",index=True,index_label="id")
#     fb_image_df.to_sql(filename_fb_ig_snapchat,con=db.engine,if_exists="append",index=True,index_label="id")
#     ig_video_df.to_sql(filename_fb_ig_snapchat,con=db.engine,if_exists="append",index=True,index_label="id")
#     ig_image_df.to_sql(filename_fb_ig_snapchat,con=db.engine,if_exists="append",index=True,index_label="id")
#     gdn_df.to_sql(filename_youtube_gdn,con=db.engine,if_exists="replace",index=True,index_label="id")
#     trueview_df.to_sql(filename_youtube_gdn,con=db.engine,if_exists="append",index=True,index_label="id")
    
#     # Storing the raw data in the Database 
#     # Calculating Daily (CPV_TRUEVIEW) and assigning type so we can separate every field 
    
#     trueview_df_daily = pd.pivot_table(trueview_df, index=['Date'],aggfunc='sum')
#     trueview_df_daily = SettingCTR_TRUEVIEW(trueview_df_daily,"trueview",trueview_budget,mvkpi_trueview)
#     trueview_df_daily_with_type = trueview_df_daily.assign(type = "trueview")
    
    
#     gdn_df_daily = pd.pivot_table(gdn_df, index=['Date'],aggfunc='sum')
#     gdn_df_daily = SettingCTR_TRUEVIEW(gdn_df_daily,"gdn",gdn_budget,mvkpi_gdn)
#     gdn_df_daily_with_type = gdn_df_daily.assign(type = "gdn")

    
    
#     # Calculating Daily (IG_IMAGE)
#     ig_image_df_daily = pd.pivot_table(ig_image_df, index=['Reporting starts'],aggfunc='sum')
#     ig_image_df_daily = SettingCTR_TRUEVIEW(ig_image_df_daily,"ig_image",ig_image_budget,mvkpi_ig_image)
#     ig_image_df_daily_with_type = ig_image_df_daily.assign(type = "ig_image")
    
#     # Calculating Daily (IG_VIDEO)
#     try:
#         ig_video_df_daily = pd.pivot_table(ig_video_df, index=['Reporting starts'],aggfunc='sum')
#         ig_video_df_daily = SettingCTR_TRUEVIEW(ig_video_df_daily,"ig_video",ig_video_budget,mvkpi_ig_video)
#         ig_video_df_daily_with_type = ig_video_df_daily.assign(type = "ig_video")
#     except:
#          pass
    
    
#     fb_image_df_daily = pd.pivot_table(fb_image_df, index=['Reporting starts'],aggfunc='sum')
#     fb_image_df_daily = SettingCTR_TRUEVIEW(fb_image_df_daily,"fb_image",fb_image_budget,mvkpi_fb_image)
#     fb_image_df_daily_with_type = fb_image_df_daily.assign(type = "fb_image")
    
    
#     fb_video_df_daily = pd.pivot_table(fb_video_df, index=['Reporting starts'],aggfunc='sum')
#     fb_video_df_daily = SettingCTR_TRUEVIEW(fb_video_df_daily,"fb_video",fb_video_budget,mvkpi_fb_video)
#     fb_video_df_daily_with_type = fb_video_df_daily.assign(type = "fb_video")

    
#     # print(trueview_df)
    
#     # Now we are storing the filename for Daily so we can store the data in database with that name 
#     filenameTotal_youtube_gdn_daily = filename_youtube_gdn + "-total" 
#     filenameTotal_fb_ig_snapchat_daily = filename_fb_ig_snapchat + "-total" 
#     # We are opening the file to write the data in the table 
#     with pd.ExcelWriter(filename+'.xlsx') as writer: 
#         # try:
#         gdn_df_daily_with_type.to_sql(filenameTotal_youtube_gdn_daily,con=db.engine,if_exists="replace",index=False)
#         trueview_df_daily_with_type.to_sql(filenameTotal_youtube_gdn_daily,con=db.engine,if_exists="append",index=False)
#         # trueview_df_daily_with_type.to_excel(workbook,sheet_name="test")
#         fb_video_df_daily_with_type.to_sql(filenameTotal_fb_ig_snapchat_daily,con=db.engine,if_exists="replace",index=False)
#         fb_image_df_daily_with_type.to_sql(filenameTotal_fb_ig_snapchat_daily,con=db.engine,if_exists="append",index=False)
#         print("1st Error here ")
#         # ig_video_df_daily_with_type.to_sql(filenameTotal_fb_ig_snapchat_daily,con=db.engine,if_exists="append",index=False)
#         print("1st Error here ")
#         ig_image_df_daily_with_type.to_sql(filenameTotal_fb_ig_snapchat_daily,con=db.engine,if_exists="append",index=False)
#         filenameTotal_fb_ig_snapchat_weekly = filename_fb_ig_snapchat + "-weekly" 
#         filenameTotal_youtube_gdn_weekly = filename_youtube_gdn + "-weekly" 
#         # We are entering the data for daily in the database 
#         # Next step is to calculate the weekly from the daily tables for each category 
#         try:
#             gdn_df_weekly=CalculatingWeekly(gdn_df_daily,"gdn")
#             # --------------------------------------------------------
#             gdn_df_weekly = gdn_df_weekly.assign(type = "gdn")
#             gdn_df_weekly.to_sql(filenameTotal_youtube_gdn_weekly,con=db.engine,if_exists="replace",index=False)
#             gdn_df_weekly.to_excel(writer, sheet_name='gdn', index=True)
            
            
#             # gdn_df_weekly.to_excel(workbook, sheet_name='gdn', index=False, startrow=1, startcol=1)
#             gdn_df_daily.to_excel(writer, sheet_name='gdn',startrow=writer.sheets["gdn"].max_row+1, index=True)
#             writer = StylingSheets(writer, "gdn")
#         except:
#             print("ERROR HERE")
#         try:
#             trueview_df_weekly=CalculatingWeekly(trueview_df_daily,"trueview")
#             # --------------------------------------------------------
#             trueview_df_weekly = trueview_df_weekly.assign(type = "trueview")
#             trueview_df_weekly.to_sql(filenameTotal_youtube_gdn_weekly,con=db.engine,if_exists="append",index=False)
#             # trueview_df_weekly.to_excel(workbook, sheet_name='trueview', index=False, startrow=1, startcol=1)
#             trueview_df_weekly.to_excel(writer, sheet_name='trueview', index=True)
#             trueview_df_daily.to_excel(writer, sheet_name='trueview',startrow=writer.sheets["trueview"].max_row+1, index=True)
#             writer = StylingSheets(writer, "trueview")
#         except:
#             pass
        
#         try:
#             fb_video_df_weekly=CalculatingWeekly(fb_video_df_daily,"fb_video")
#             # --------------------------------------------------------
#             fb_video_df_weekly = fb_video_df_weekly.assign(type = "fb_video")
#             fb_video_df_weekly.to_sql(filenameTotal_fb_ig_snapchat_weekly,con=db.engine,if_exists="replace",index=True,index_label="id")
#             fb_video_df_weekly.to_excel(writer, sheet_name='fb_video', index=True)
#             fb_video_df_daily.to_excel(writer, sheet_name='fb_video',startrow=writer.sheets["fb_video"].max_row+1, index=True)
#             writer = StylingSheets(writer, "fb_video")
#             # fb_video_df_weekly.to_excel(workbook, sheet_name='fb_video', index=False, startrow=1, startcol=1)
#         except:
#             print("Any Error")
#         try:
#             fb_image_df_weekly=CalculatingWeekly(fb_image_df_daily,"fb_image")
#             # --------------------------------------------------------
#             fb_image_df_weekly = fb_image_df_weekly.assign(type = "fb_image")
#             fb_image_df_weekly.to_sql(filenameTotal_fb_ig_snapchat_weekly,con=db.engine,if_exists="append",index=True,index_label="id")
#             # fb_image_df_weekly.to_excel(workbook, sheet_name='fb_image', index=False, startrow=1, startcol=1)
#             fb_image_df_weekly.to_excel(writer, sheet_name='fb_image', index=True)
#             fb_image_df_daily.to_excel(writer, sheet_name='fb_image',startrow=writer.sheets["fb_image"].max_row+1, index=True)
#             writer = StylingSheets(writer, "fb_image")
#         except:
#             pass
        
        
#         try:
#             ig_image_df_weekly=CalculatingWeekly(ig_image_df_daily,"ig_image")
#             # --------------------------------------------------------
#             ig_image_df_weekly = ig_image_df_weekly.assign(type = "ig_image")
#             ig_image_df_weekly.to_sql(filenameTotal_fb_ig_snapchat_weekly,con=db.engine,if_exists="append",index=True,index_label="id")
#             ig_image_df_weekly.to_excel(writer, sheet_name='ig_image', index=True)
#             ig_image_df_daily.to_excel(writer, sheet_name='ig_image',startrow=writer.sheets["ig_image"].max_row+1, index=True)
#             writer = StylingSheets(writer, "ig_image")
            
#             # ig_image_df_weekly.to_excel(workbook, sheet_name='ig_image', index=False, startrow=1, startcol=1)
#         except:
#             pass
#         try:
#             ig_video_df_weekly=CalculatingWeekly(ig_video_df_daily,"ig_video")
#             # --------------------------------------------------------
#             ig_video_df_weekly = ig_video_df_weekly.assign(type = "ig_video")
#             ig_video_df_weekly.to_sql(filenameTotal_fb_ig_snapchat_weekly,con=db.engine,if_exists="append",index=True,index_label="id")
#             ig_video_df_weekly.to_excel(writer, sheet_name='ig_video', index=True)
#             ig_video_df_daily.to_excel(writer, sheet_name='ig_video',startrow=writer.sheets["ig_video"].max_row+1, index=True)
#             try:
#                 writer = StylingSheets(writer, "ig_video")
#             except:
#                 pass
            
#             # Its not working make it work
#             # ig_video_df_weekly.to_excel(workbook, sheet_name='ig_video', index=False, startrow=1, startcol=1)
#         except:
#             pass
        
#         # Till here we have calculated weekly data from daily and stored in weekly table 
#         # Now we will calculate the weekly total for all the weeks 
#         try:
#             ig_image_df_weekly_total = Weekly_Total(ig_image_df_weekly,mvkpi_ig_image,ig_image_budget,"ig_image")
#             # ig_video_df_weekly_total = Weekly_Total(ig_video_df_weekly,mvkpi_ig_video,ig_video_budget)
#             fb_video_df_weekly_total = Weekly_Total(fb_video_df_weekly,mvkpi_fb_video,fb_video_budget,"fb_video")
#             fb_image_df_weekly_total = Weekly_Total(fb_image_df_weekly,mvkpi_fb_image,fb_image_budget,"fb_image")
#             trueview_df_weekly_total = Weekly_Total(trueview_df_weekly,mvkpi_trueview,trueview_budget,"trueview")
#             gdn_df_weekly_total = Weekly_Total(gdn_df_weekly,mvkpi_gdn,gdn_budget,"gdn")
            
#             # print(fb_video['Clicks'])
#             print("Filename")
#             weekly_total = pd.concat([ig_image_df_weekly_total,
#             # ig_video_df_weekly_total,
#             fb_video_df_weekly_total,fb_image_df_weekly_total,trueview_df_weekly_total,gdn_df_weekly_total])
#             weekly_total.to_sql(filename+"-weekly_total",con=db.engine,if_exists="replace",index=True)
#             # print(filename)
#             # weekly_total.to_excel(writer, sheet_name='summary_table', index=False)
#         except:
#             print("ERROR")
#         # Now once we have weekly total we are now going for Daily Total Tab 
#         try:
#             # Adding all the total tables together to get a daily table with record 
#             new_df = pd.concat([trueview_df_daily,gdn_df_daily,fb_image_df_daily,fb_video_df_daily,ig_image_df_daily,ig_video_df_daily])
#             # Doing sum ...
#             new_df = pd.pivot_table(new_df, index=new_df.index,aggfunc='sum')
#             # Daily Report Total 
#             new_df['CTR'] = round((new_df['Clicks'] / new_df['Impressions'])*100,2)
#             # Daily Report CTR 
            
#             # Adding mvkpi for all inputs 
#             totalMVKPI = int(mvkpi_fb_image) + int(mvkpi_fb_video) + int(mvkpi_gdn) + int(mvkpi_ig_image) +int(mvkpi_ig_video)+ int(mvkpi_trueview)  
            
#             # Adding budget for all inputs 
#             totalBudget = int(gdn_budget) + int(trueview_budget) + int(fb_video_budget) + int(fb_image_budget) + int(ig_image_budget) + int(ig_video_budget)
            
#             # Total Monthly View KPI 
#             new_df['DAILY_VIEW_KPI'] = round(totalMVKPI / 30)
#             new_df['Budget'] = round(new_df['Budget'])
#             new_df['SPENT_BUDGET'] = round((new_df['Budget']/totalBudget)*100)
#             new_df = new_df.assign(Views_YT_FB_IG_GDN = new_df['TrueView: Views'] + new_df['3 Views'])
#             new_df["DAILY_KPI_ACHIEVEMENT"] = round(new_df["Views_YT_FB_IG_GDN"]/ new_df["DAILY_VIEW_KPI"]*100)
#             new_df = new_df.assign(CPC = round(new_df['Budget'] / new_df['Clicks'],2))
            
#             # date_rng = pd.date_range(start=new_df.index.min(), end=new_df.index.max(), freq='D')
         
        
#             # for i in range(0, len(date_rng)):
                
#             #     row = new_df.iloc[i].select_dtypes(include=['int64','double'])
#             #     new_df.iloc[i]['CTR'] = round((new_df.iloc[i]['Clicks'] / new_df.iloc[i]['Impressions'])*100),2
        
#             new_df.to_sql(filename+"-daily_total",con=db.engine,if_exists="replace",index=True)
#             # summary_df = new_df.sum(axis=1)
#             # summary_df = new_df.sum()
#             # print(summary_df)
#             # summary_df = new_df.iloc[0:-1].select_dtypes(include=['double']).sum()
#             # So Daily Report is correct 
#             # Now we will move to summary 
            
#             summary = new_df.copy()
        
        
        
#             summary.drop(summary.index, inplace=True)
#             # print(gdn_df_daily)
#             for i in range(0, 1):
#                 row = new_df.iloc[i:len(new_df)].select_dtypes(include=['int64','double']).sum()
#                 row['CTR'] = row['Clicks'] / row['Impressions']
#                 # print(totalBudget)
#                 row['Budget'] = round(row['Budget'],2)
#                 row['SPENT_BUDGET'] = round((row['Budget']/totalBudget)*100)
#                 row['CTR'] = round(row['CTR']*100,2)
#                 row['3 Views'] = row["Views_YT_FB_IG_GDN"]
#                 # row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
#                 # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
#                 summary = summary.append(row, ignore_index=True)
#             #
            
#             summary = summary.assign(MVKPI_Total = totalMVKPI)
#             summary.to_sql("summary",con=db.engine,if_exists="replace",index=False)
#             # summary.to_excel(writer, sheet_name='summary', index=False)
#             # summary.to_excel(writer,con=db.engine,if_exist="replace",index=False)
            
        
#             # new_df.set_index('Date', inplace=True)
#             new_df.to_excel(writer, sheet_name='daily_report', index=True,index_label="Date")
#             writer = StylingSheets(writer, "daily_report")
            
#         except:
#             print("Didn't matched")
#         # except:
#         #     print("This is running")
#         #     db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_fb_ig_snapchat_weekly}`")
#         #     db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_youtube_gdn_weekly}`")
#         #     db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_youtube_gdn_daily}`") 
#         #     db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_fb_ig_snapchat_daily}`") 
#         #     db.engine.execute(f"DROP table IF EXISTS `{filename_youtube_gdn}`") 
#         #     db.engine.execute(f"DROP table IF EXISTS `{filename_fb_ig_snapchat}`") 
        

#         # start_date = gdn_df.index.min()
#         # end_date = gdn_df.index.max()
        
#     # except:
#     #     print("There is error")
#     #     db.engine.execute(f"DELETE FROM media_files WHERE filename='{filename}'")
#     #     db.engine.execute(f"DROP table IF EXISTS `{filename_youtube_gdn}`") 
#     #     db.engine.execute(f"DROP table IF EXISTS `{filename_fb_ig_snapchat}`") 

#     #     filenameTotal_youtube_gdn_daily = filename_youtube_gdn + "-total" 
#     #     filenameTotal_fb_ig_snapchat_daily = filename_fb_ig_snapchat + "-total" 
#     #     try:
#     #         gdn_df_daily.to_sql(filenameTotal_youtube_gdn_daily,con=db.engine,if_exists="fail",index=True,index_label="id")
#     #         db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_youtube_gdn_daily}`") 
            
#     #         fb_image_df_daily.to_sql(filenameTotal_fb_ig_snapchat_daily,con=db.engine,if_exists="fail",index=True,index_label="id")
#     #         db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_fb_ig_snapchat_daily}`") 
#     #     except:
#     #         db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_youtube_gdn_daily}`") 
#     #         db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_fb_ig_snapchat_daily}`") 


#     #     filenameTotal_youtube_gdn_weekly = filename_youtube_gdn + "-weekly"
#     #     filenameTotal_fb_ig_snapchat_weekly = filename_fb_ig_snapchat + "-weekly"
#     #     try:
#     #         db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_youtube_gdn_weekly}`")
#     #         db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_fb_ig_snapchat_weekly}`")
#     #     except:
#     #         pass
#         # Now Storing total in the database 
    
#     keys = ['Impressions','Clicks','Budget','Video watches at 25%', 'Video watches at 50%', 'Video watches at 75%', 'View 100%']
    
#     fb_video_data = DataForSummary(filename,'fb_video')
#     fb_image_data = DataForSummary(filename,'fb_image')
#     ig_video_data = DataForSummary(filename,'ig_video')
#     ig_image_data = DataForSummary(filename,'ig_image')
#     trueview_data = DataForSummary(filename,'trueview')
#     gdn_data = DataForSummary(filename,'gdn')

#     summary_data = DataForSummaryTable()

#     thin_border = Border(left=Side(style='thin'), 
#     right=Side(style='thin'), 
#     top=Side(style='thin'), 
#     bottom=Side(style='thin'))
#     try:
#         default_template = load_workbook(filename+".xlsx")
#         sheet = default_template.create_sheet("Summary")

        
#         starting_char = "F"
#         starting_index = 10

#         # Loop for Characters 
#             # Loop for numbers 
#         list_of_vertical_headings = ["FB VIDEO","FB IMAGE","IG IMAGE","TRUEVIEW","GDN","TOTAL" ]
#         list_of_horizontal_headings = ["VIEWS","MONTHLY KPI ACHIEVEMENT","CLICKS","CTR","BUDGET","BUDGET SPENT" ]
#         for i in range(0,7): 
#             if i == 0:
                
#                 ind = 0 
#                 for heading in list_of_vertical_headings:
                    
#                     sheet[chr(ord(starting_char) + i) + str(starting_index+ind)] = heading 
#                     StylingSummary(chr(ord(starting_char) + i) + str(starting_index+ind),sheet,type="heading")
#                     ind +=1 

            

#         sheet["G9"] = "VIEWS" 
#         StylingSummary("G9",sheet,type="heading")
        
        
#         sheet["G10"] = fb_video_data['VIEWS']
#         StylingSummary("G10",sheet)

#         sheet["G11"] = fb_image_data['VIEWS']
#         StylingSummary("G11",sheet)

#         # sheet["G12"] = ig_video_data['VIEWS']
#         sheet["G12"] = ig_image_data['VIEWS']
#         StylingSummary("G12",sheet)

#         sheet["G13"] = trueview_data['VIEWS']
#         StylingSummary("G13",sheet)

#         sheet["G14"] = gdn_data['VIEWS']
#         StylingSummary("G14",sheet)

#         sheet["G15"] = summary_data["VIEWS"]
#         StylingSummary("G15",sheet)

#         sheet["H9"] = "MONTHLY KPI ACHIEVEMENT"
#         StylingSummary("H9",sheet,type="heading")

#         sheet["H10"] = fb_video_data['MONTHLY_KPI_ACHIEVEMENT']
#         StylingSummary("H10",sheet)

#         sheet["H11"] = fb_image_data['MONTHLY_KPI_ACHIEVEMENT']
#         StylingSummary("H11",sheet)

#         # sheetH"G12"] = ig_video_data['MONTHLY_KPI_ACHIEVEMENT']
#         sheet["H12"] = ig_image_data['MONTHLY_KPI_ACHIEVEMENT']
#         StylingSummary("H12",sheet)

#         sheet["H13"] = trueview_data['MONTHLY_KPI_ACHIEVEMENT']
#         StylingSummary("H13",sheet)

#         sheet["H14"] = gdn_data['MONTHLY_KPI_ACHIEVEMENT']
#         StylingSummary("H14",sheet)
        
#         sheet["H15"] = summary_data["MONTHLY_KPI_ACHIEVEMENT"]
#         StylingSummary("H15",sheet)

#         sheet['I9'] = "CLICKS"
#         StylingSummary("I9",sheet,type="heading")

#         sheet["I10"] = fb_video_data['CLICKS']
#         StylingSummary("I10",sheet)

#         sheet["I11"] = fb_image_data['CLICKS']
#         StylingSummary("I11",sheet)
#         # sheetI"G12"] = ig_video_data['CLICKS']
#         sheet["I12"] = ig_image_data['CLICKS']
#         StylingSummary("I12",sheet)

#         sheet["I13"] = trueview_data['CLICKS']
#         StylingSummary("I13",sheet)

#         sheet["I14"] = gdn_data['CLICKS']
#         StylingSummary("I14",sheet)

#         sheet["I15"] = summary_data["CLICKS"]
#         StylingSummary("I15",sheet)

#         sheet["J9"] = "CTR"
#         StylingSummary("J9",sheet,type="heading")

#         sheet["J10"] = fb_video_data['CTR']
#         StylingSummary("J10",sheet)

#         sheet["J11"] = fb_image_data['CTR']
#         StylingSummary("J11",sheet)

#         # sheetJ"G12"] = ig_video_data['CTR']
#         sheet["J12"] = ig_image_data['CTR']
#         StylingSummary("J12",sheet)
#         sheet["J13"] = trueview_data['CTR']
#         StylingSummary("J13",sheet)
#         sheet["J14"] = gdn_data['CTR']
#         StylingSummary("J14",sheet)
#         sheet["J15"] = summary_data["CTR"]
#         StylingSummary("J15",sheet)

#         sheet["K9"] = "BUDGET"
#         StylingSummary("K9",sheet,type="heading")

#         sheet["K10"] = fb_video_data['BUDGET']
#         StylingSummary("K10",sheet)

#         sheet["K11"] = fb_image_data['BUDGET']
#         StylingSummary("K11",sheet)
#         # sheetK"G12"] = ig_video_data['BUDGET']
#         sheet["K12"] = ig_image_data['BUDGET']
#         StylingSummary("K12",sheet)
#         sheet["K13"] = trueview_data['BUDGET']
#         StylingSummary("K13",sheet)
#         sheet["K14"] = gdn_data['BUDGET']
#         StylingSummary("K14",sheet)
#         sheet["K15"] = summary_data["BUDGET"]
#         StylingSummary("K15",sheet)

#         sheet["L9"] = "BUDGET SPENT"
#         StylingSummary("L9",sheet,type="heading")
#         sheet["L10"] = fb_video_data['SPENT_BUDGET']
#         StylingSummary("L10",sheet)
#         sheet["L11"] = fb_image_data['SPENT_BUDGET']
#         StylingSummary("L11",sheet)
#         # sheetL"G12"] = ig_video_data['SPENT_BUDGET']
#         sheet["L12"] = ig_image_data['SPENT_BUDGET']
#         StylingSummary("L12",sheet)
#         sheet["L13"] = trueview_data['SPENT_BUDGET']
#         StylingSummary("L13",sheet)
#         sheet["L14"] = gdn_data['SPENT_BUDGET']
#         StylingSummary("L14",sheet)
#         sheet["L15"] = summary_data["SPENT_BUDGET"]
#         StylingSummary("L15",sheet)

#         default_template.save(filename+".xlsx")

#         serialized_workbook = pickle.dumps(default_template)
#         uploading_file = media_files(filename,serialized_workbook,current_user.username)
#         db.session.add(uploading_file)
#         db.session.commit()
#         # workbook.save("summary.xlsx")
#         # print("Runn")
#         # os.remove(filename+".xlsx")
#     except:
#         pass

#     # Save the file
#     # workbook.save("weekly_report.xlsx")



#     return jsonify({"message":"Uploaded successfully"}),200




# # New Campaign 
# @app.route("/raw_data_new_campaign",methods=["POST"])
# @cross_origin()
# @token_required

# def raw_data_new_campaign(current_user):

#     filename = request.form["filename"]
#     # File name to be 
#     fb_video = request.form['fb_video']
#     fb_video_budget = request.form["fb_video_budget"]

#     ig_video = request.form['ig_video']
#     ig_video_budget = request.form["ig_video_budget"]
    
#     sc_video = request.form['sc_video']
#     sc_video_budget = request.form["sc_video_budget"]

#     twitter_video = request.form['twitter_video']
#     twitter_video_budget = request.form["twitter_video_budget"]

#     dv360 = request.form['dv360']
#     dv360_budget = request.form["dv360_budget"]
    
   
#     mvkpi_dv360 = request.form['mvkpi_dv360']
#     mvkpi_fb_video = request.form['mvkpi_fb_video']
#     mvkpi_ig_video = request.form['mvkpi_ig_video']
#     mvkpi_sc_video = request.form['mvkpi_sc_video']
#     mvkpi_twitter_video = request.form['mvkpi_twitter_video']

#     raw_data_file_fb_ig_snapchat_twitter = request.files["file-fb-ig-snapchat_twitter"]
#     raw_data_file_dv360 = request.files["file-dv360"]


#     # Taking the files and the data as Input and saving in variable 


#     # We are getting the files 


#     df_fb_ig_snapchat_twitter = pd.read_excel(raw_data_file_fb_ig_snapchat_twitter)
#     df_dv360 = pd.read_excel(raw_data_file_dv360)

#     # We are reading the excel files 


#     # Getting dataframe for each date. 
#     # Getting status that either the data is stored or not 
#     # If data is stored then get the sum of the values 
#     # And store the calculated data in the database table -- ctr, and other calcualtion of the table 
#     # 
#     filename_dv360 = filename + "-dv360" 
#     filename_fb_ig_snapchat_twitter = filename + "-fb_ig_snapchat_twitter" 
#     # Stored file names of two different categories

#     # We are adding Dataframes to SQL Table 
#     fb_video_df = DataFrameToSQL(df_fb_ig_snapchat_twitter,fb_video,"fb_video")
    
#     ig_video_df = DataFrameToSQL(df_fb_ig_snapchat_twitter,ig_video,"ig_video")

#     sc_video_df = DataFrameToSQL(df_fb_ig_snapchat_twitter,sc_video,"sc_video")
    
#     twitter_video_df = DataFrameToSQL(df_fb_ig_snapchat_twitter,twitter_video,"twitter_video")

#     # -------------------------------------------------------------
#     dv360_df = DataFrameToSQL(df_dv360,dv360,"dv360")
    
#     # Assigning Types to raw data and returning the dataframe to be stored in SQL Table

#     # workbook = Workbook()
#     # trueview_df.to_excel(workbook,"trueview_test",index=False)
#     try:
#     # Raw data stored in the raw data table 
#         fb_video_df.to_sql(filename_fb_ig_snapchat_twitter,con=db.engine,if_exists="fail",index=True,index_label="id")
#         ig_video_df.to_sql(filename_fb_ig_snapchat_twitter,con=db.engine,if_exists="append",index=True,index_label="id")
#         sc_video_df.to_sql(filename_fb_ig_snapchat_twitter,con=db.engine,if_exists="append",index=True,index_label="id")
#         twitter_video_df.to_sql(filename_fb_ig_snapchat_twitter,con=db.engine,if_exists="append",index=True,index_label="id")
#         dv360_df.to_sql(filename_dv360,con=db.engine,if_exists="fail",index=True,index_label="id")
        
        
#         # Storing the raw data in the Database 
#         # Calculating Daily (CPV_TRUEVIEW) and assigning type so we can separate every field 
        

        
#         dv360_df_daily = pd.pivot_table(dv360_df, index=['Date'],aggfunc='sum')
#         dv360_df_daily = SettingCTR_TRUEVIEW(dv360_df_daily,"dv360",dv360_budget,mvkpi_dv360)
#         dv360_df_daily_with_type = dv360_df_daily.assign(type = "dv360")

        
        
#         # Calculating Daily (IG_IMAGE)
        
#         # Calculating Daily (IG_VIDEO)
#         ig_video_df_daily = pd.pivot_table(ig_video_df, index=['Reporting starts'],aggfunc='sum')
#         ig_video_df_daily = SettingCTR_TRUEVIEW(ig_video_df_daily,"ig_video",ig_video_budget,mvkpi_ig_video)
#         ig_video_df_daily_with_type = ig_video_df_daily.assign(type = "ig_video")
        
#         fb_video_df_daily = pd.pivot_table(fb_video_df, index=['Reporting starts'],aggfunc='sum')
#         fb_video_df_daily = SettingCTR_TRUEVIEW(fb_video_df_daily,"fb_video",fb_video_budget,mvkpi_fb_video)
#         fb_video_df_daily_with_type = fb_video_df_daily.assign(type = "fb_video")

#         sc_video_df_daily = pd.pivot_table(sc_video_df, index=['Reporting starts'],aggfunc='sum')
#         sc_video_df_daily = SettingCTR_TRUEVIEW(sc_video_df_daily,"sc_video",sc_video_budget,mvkpi_sc_video)
#         sc_video_df_daily_with_type = sc_video_df_daily.assign(type = "sc_video")

#         twitter_video_df_daily = pd.pivot_table(twitter_video_df, index=['Reporting starts'],aggfunc='sum')
#         twitter_video_df_daily = SettingCTR_TRUEVIEW(twitter_video_df_daily,"twitter_video",twitter_video_budget,mvkpi_twitter_video)
#         twitter_video_df_daily_with_type = twitter_video_df_daily.assign(type = "twitter_video")
#         # print(trueview_df)
        
#         # Now we are storing the filename for Daily so we can store the data in database with that name 
#         filenameTotal_dv360_daily = filename_dv360 + "-total" 
#         filenameTotal_fb_ig_snapchat_twitter_daily = filename_fb_ig_snapchat_twitter + "-total" 
#         # We are opening the file to write the data in the table 
#         with pd.ExcelWriter(filename+'.xlsx') as writer: 
#             try:
#                 dv360_df_daily_with_type.to_sql(filenameTotal_dv360_daily,con=db.engine,if_exists="replace",index=False)
                
#                 # trueview_df_daily_with_type.to_excel(workbook,sheet_name="test")
#                 fb_video_df_daily_with_type.to_sql(filenameTotal_fb_ig_snapchat_twitter_daily,con=db.engine,if_exists="replace",index=False)
#                 ig_video_df_daily_with_type.to_sql(filenameTotal_fb_ig_snapchat_twitter_daily,con=db.engine,if_exists="append",index=False)
#                 sc_video_df_daily_with_type.to_sql(filenameTotal_fb_ig_snapchat_twitter_daily,con=db.engine,if_exists="append",index=False)
#                 twitter_video_df_daily_with_type.to_sql(filenameTotal_fb_ig_snapchat_twitter_daily,con=db.engine,if_exists="append",index=False)
#                 filenameTotal_fb_ig_snapchat_twitter_weekly = filename_fb_ig_snapchat_twitter + "-weekly" 
#                 filenameTotal_dv360_weekly = filename_dv360 + "-weekly" 
#                 # We are entering the data for daily in the database 
#                 # Next step is to calculate the weekly from the daily tables for each category 
#                 try:
#                     dv360_df_weekly=CalculatingWeekly(dv360_df_daily,"dv360")
#                     # --------------------------------------------------------
#                     dv360_df_weekly = dv360_df_weekly.assign(type = "dv360")
#                     dv360_df_weekly.to_sql(filenameTotal_dv360_weekly,con=db.engine,if_exists="replace",index=True,index_label="id")
#                     dv360_df_weekly.to_excel(writer, sheet_name='dv360', index=True)
                    
                    
#                     # gdn_df_weekly.to_excel(workbook, sheet_name='gdn', index=False, startrow=1, startcol=1)
#                     dv360_df_daily.to_excel(writer, sheet_name='dv360',startrow=writer.sheets["dv360"].max_row+1, index=True)
#                     writer = StylingSheets(writer, "dv360")
#                 except:
#                     print("ERROR HERE")
            
#                 try:
#                     fb_video_df_weekly=CalculatingWeekly(fb_video_df_daily,"fb_video")
#                     # --------------------------------------------------------
#                     fb_video_df_weekly = fb_video_df_weekly.assign(type = "fb_video")
#                     fb_video_df_weekly.to_sql(filenameTotal_fb_ig_snapchat_twitter_weekly,con=db.engine,if_exists="replace",index=True,index_label="id")
#                     fb_video_df_weekly.to_excel(writer, sheet_name='fb_video', index=True)
#                     fb_video_df_daily.to_excel(writer, sheet_name='fb_video',startrow=writer.sheets["fb_video"].max_row+1, index=True)
#                     writer = StylingSheets(writer, "fb_video")
#                     # fb_video_df_weekly.to_excel(workbook, sheet_name='fb_video', index=False, startrow=1, startcol=1)
#                 except:
#                     print("Any Error")
#                 try:
#                     sc_video_df_weekly=CalculatingWeekly(sc_video_df_daily,"sc_video")
#                     # --------------------------------------------------------
#                     sc_video_df_weekly = sc_video_df_weekly.assign(type = "sc_video")
#                     sc_video_df_weekly.to_sql(filenameTotal_fb_ig_snapchat_twitter_weekly,con=db.engine,if_exists="append",index=True,index_label="id")
#                     sc_video_df_weekly.to_excel(writer, sheet_name='sc_video', index=True)
#                     sc_video_df_daily.to_excel(writer, sheet_name='sc_video',startrow=writer.sheets["sc_video"].max_row+1, index=True)
#                     try:
#                         writer = StylingSheets(writer, "sc_video")
#                     except:
#                         pass
                    
#                     # Its not working make it work
#                     # ig_video_df_weekly.to_excel(workbook, sheet_name='ig_video', index=False, startrow=1, startcol=1)
#                 except:
#                     pass
                
#                 try:
#                     twitter_video_df_weekly=CalculatingWeekly(twitter_video_df_daily,"twitter_video")
#                     # --------------------------------------------------------
#                     twitter_video_df_weekly = twitter_video_df_weekly.assign(type = "twitter_video")
#                     twitter_video_df_weekly.to_sql(filenameTotal_fb_ig_snapchat_twitter_weekly,con=db.engine,if_exists="append",index=True,index_label="id")
#                     twitter_video_df_weekly.to_excel(writer, sheet_name='twitter_video', index=True)
#                     twitter_video_df_daily.to_excel(writer, sheet_name='twitter_video',startrow=writer.sheets["twitter_video"].max_row+1, index=True)
#                     try:
#                         writer = StylingSheets(writer, "twitter_video")
#                     except:
#                         pass
                    
#                     # Its not working make it work
#                     # ig_video_df_weekly.to_excel(workbook, sheet_name='ig_video', index=False, startrow=1, startcol=1)
#                 except:
#                     pass
#                 try:
#                     ig_video_df_weekly=CalculatingWeekly(ig_video_df_daily,"ig_video")
#                     # --------------------------------------------------------
#                     ig_video_df_weekly = ig_video_df_weekly.assign(type = "ig_video")
#                     ig_video_df_weekly.to_sql(filenameTotal_fb_ig_snapchat_twitter_weekly,con=db.engine,if_exists="append",index=True,index_label="id")
#                     ig_video_df_weekly.to_excel(writer, sheet_name='ig_video', index=True)
#                     ig_video_df_daily.to_excel(writer, sheet_name='ig_video',startrow=writer.sheets["ig_video"].max_row+1, index=True)
#                     try:
#                         writer = StylingSheets(writer, "ig_video")
#                     except:
#                         pass
                    
#                     # Its not working make it work
#                     # ig_video_df_weekly.to_excel(workbook, sheet_name='ig_video', index=False, startrow=1, startcol=1)
#                 except:
#                     pass

#                 try:
#                         ig_video_df_weekly_total = Weekly_Total(ig_video_df_weekly,mvkpi_ig_video,ig_video_budget,"ig_video")
#                         fb_video_df_weekly_total = Weekly_Total(fb_video_df_weekly,mvkpi_fb_video,fb_video_budget,"fb_video")
#                         sc_video_df_weekly_total = Weekly_Total(sc_video_df_weekly,mvkpi_sc_video,sc_video_budget,"sc_video")
#                         twitter_video_df_weekly_total = Weekly_Total(twitter_video_df_weekly,mvkpi_twitter_video,twitter_video_budget,"twitter_video")
#                         dv360_df_weekly_total = Weekly_Total(dv360_df_weekly,mvkpi_dv360,dv360_budget,"dv360")
                        
#                         # print(fb_video['Clicks'])

#                         print("Filename")
#                         weekly_total = pd.concat([ig_video_df_weekly_total,
#                         # ig_video_df_weekly_total,
#                         fb_video_df_weekly_total,sc_video_df_weekly_total,twitter_video_df_weekly_total,dv360_df_weekly_total])
#                         weekly_total.to_sql(filename+"-weekly_total",con=db.engine,if_exists="replace",index=True)
#                         # print(filename)
#                         # weekly_total.to_excel(writer, sheet_name='summary_table', index=False)


#                 except:
#                     print("ERROR")
                
#                 # Till here we have calculated weekly data from daily and stored in weekly table 
#                 # Now we will calculate the weekly total for all the weeks 
            
#                 # Now once we have weekly total we are now going for Daily Total Tab 
#                 try:
#                     # Adding all the total tables together to get a daily table with record 
#                     new_df = pd.concat([dv360_df_daily,fb_video_df_daily,ig_video_df_daily,sc_video_df_daily,twitter_video_df_daily])
#                     # Doing sum ...
#                     print(new_df)
#                     new_df = pd.pivot_table(new_df, index=new_df.index,aggfunc='sum')
#                     # Daily Report Total 
#                     new_df['CTR'] = round((new_df['Clicks'] / new_df['Impressions'])*100,2)
#                     # Daily Report CTR 
                    
#                     # Adding mvkpi for all inputs 
#                     totalMVKPI = int(mvkpi_sc_video) + int(mvkpi_fb_video) + int(mvkpi_dv360) + int(mvkpi_twitter_video) +int(mvkpi_ig_video)
                    
#                     # Adding budget for all inputs 
#                     totalBudget = int(dv360_budget)  + int(fb_video_budget) + int(ig_video_budget)+ int(sc_video_budget)+ int(twitter_video_budget)
                    
#                     # Total Monthly View KPI 
                    
#                     new_df['DAILY_VIEW_KPI'] = round(totalMVKPI / 30)
#                     new_df['Budget'] = round(new_df['Budget'])
#                     new_df['SPENT_BUDGET'] = round((new_df['Budget']/totalBudget)*100)
#                     new_df = new_df.assign(Views_YT_FB_IG_GDN = new_df['TrueView: Views'] + new_df['3 Views'])
#                     print("Working till heree---------")
#                     new_df["DAILY_KPI_ACHIEVEMENT"] = round(new_df["Views_YT_FB_IG_GDN"]/ new_df["DAILY_VIEW_KPI"]*100)
#                     new_df = new_df.assign(CPC = round(new_df['Budget'] / new_df['Clicks'],2))
                    
#                     # date_rng = pd.date_range(start=new_df.index.min(), end=new_df.index.max(), freq='D')
                
            
#                     # for i in range(0, len(date_rng)):
                        
#                     #     row = new_df.iloc[i].select_dtypes(include=['int64','double'])
#                     #     new_df.iloc[i]['CTR'] = round((new_df.iloc[i]['Clicks'] / new_df.iloc[i]['Impressions'])*100),2
            
#                     new_df.to_sql(filename+"-daily_total",con=db.engine,if_exists="replace",index=False)
#                     # summary_df = new_df.sum(axis=1)
#                     # summary_df = new_df.sum()
#                     # print(summary_df)
#                     # summary_df = new_df.iloc[0:-1].select_dtypes(include=['double']).sum()
#                     # So Daily Report is correct 
#                     # Now we will move to summary 
                    
#                     summary = new_df.copy()
                
                
                
#                     summary.drop(summary.index, inplace=True)
#                     # print(gdn_df_daily)
#                     for i in range(0, 1):
#                         row = new_df.iloc[i:len(new_df)].select_dtypes(include=['int64','double']).sum()
#                         row['CTR'] = row['Clicks'] / row['Impressions']
#                         # print(totalBudget)
#                         row['Budget'] = round(row['Budget'],2)
#                         row['SPENT_BUDGET'] = round((row['Budget']/totalBudget)*100)
#                         row['CTR'] = round(row['CTR']*100,2)
#                         row['3 Views'] = row["Views_YT_FB_IG_GDN"]
#                         # row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
#                         # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
#                         summary = summary.append(row, ignore_index=True)
#                     #
                    
#                     summary = summary.assign(MVKPI_Total = totalMVKPI)
#                     summary.to_sql("summary",con=db.engine,if_exists="replace",index=False)
#                     # summary.to_excel(writer, sheet_name='summary', index=False)
#                     # summary.to_excel(writer,con=db.engine,if_exist="replace",index=False)
                    
            
#                     # new_df.set_index('Date', inplace=True)
#                     new_df.to_excel(writer, sheet_name='daily_report', index=True,index_label="Date")
#                     writer = StylingSheets(writer, "daily_report")
                    
#                 except:
#                     print("Didn't matched")
#             except:
#                 print("This is running")
#                 db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_fb_ig_snapchat_twitter_weekly}`")
#                 db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_dv360_weekly}`")
#                 db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_dv360_daily}`") 
#                 db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_fb_ig_snapchat_twitter_daily}`") 
#                 db.engine.execute(f"DROP table IF EXISTS `{filename_dv360}`") 
#                 db.engine.execute(f"DROP table IF EXISTS `{filename_fb_ig_snapchat_twitter}`") 
        

#         # start_date = gdn_df.index.min()
#         # end_date = gdn_df.index.max()
        
#     except:
#         print("There is error")
#         db.engine.execute(f"DELETE FROM media_files WHERE filename='{filename}'")
#         db.engine.execute(f"DROP table IF EXISTS `{filename_dv360}`") 
#         db.engine.execute(f"DROP table IF EXISTS `{filename_fb_ig_snapchat_twitter}`") 

#         filenameTotal_dv360_daily = filename_dv360+ "-total" 
#         filenameTotal_fb_ig_snapchat_twitter_daily = filename_fb_ig_snapchat_twitter + "-total" 
#         try:
#             dv360_df_daily.to_sql(filenameTotal_dv360_daily,con=db.engine,if_exists="fail",index=True,index_label="id")
#             db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_dv360_daily}`") 
            
#             fb_video_df_daily.to_sql(filenameTotal_fb_ig_snapchat_twitter_daily,con=db.engine,if_exists="fail",index=True,index_label="id")
#             db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_fb_ig_snapchat_twitter_daily}`") 
#         except:
#             db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_dv360_daily}`") 
#             db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_fb_ig_snapchat_twitter_daily}`") 


#         filenameTotal_youtube_gdn_weekly = filename_dv360 + "-weekly"
#         filenameTotal_fb_ig_snapchat_weekly = filename_fb_ig_snapchat_twitter + "-weekly"
#         try:
#             db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_youtube_gdn_weekly}`")
#             db.engine.execute(f"DROP table IF EXISTS `{filenameTotal_fb_ig_snapchat_weekly}`")
#         except:
#             pass
#         # Now Storing total in the database 
    
#     # keys = ['Impressions','Clicks','Budget','Video watches at 25%', 'Video watches at 50%', 'Video watches at 75%', 'View 100%']
    
#     fb_video_data = DataForSummary(filename,'fb_video')
#     ig_video_data = DataForSummary(filename,'ig_video')
#     sc_video_data = DataForSummary(filename,'sc_video')
#     twitter_video_data = DataForSummary(filename,'twitter_video')
#     dv360_data = DataForSummary(filename,'dv360')

#     summary_data = DataForSummaryTable()

#     thin_border = Border(left=Side(style='thin'), 
#     right=Side(style='thin'), 
#     top=Side(style='thin'), 
#     bottom=Side(style='thin'))
#     try:
#         default_template = load_workbook(filename+".xlsx")
#         sheet = default_template.create_sheet("Summary")

        
#         starting_char = "F"
#         starting_index = 10

#         # Loop for Characters 
#             # Loop for numbers 
#         list_of_vertical_headings = ["FB VIDEO","SC VIDEO","IG VIDEO","TWITTER VIDEO","DV360","TOTAL" ]
#         list_of_horizontal_headings = ["VIEWS","MONTHLY KPI ACHIEVEMENT","CLICKS","CTR","BUDGET","BUDGET SPENT" ]
#         for i in range(0,7): 
#             if i == 0:
                
#                 ind = 0 
#                 for heading in list_of_vertical_headings:
                    
#                     sheet[chr(ord(starting_char) + i) + str(starting_index+ind)] = heading 
#                     StylingSummary(chr(ord(starting_char) + i) + str(starting_index+ind),sheet,type="heading")
#                     ind +=1 

            

#         sheet["G9"] = "VIEWS" 
#         StylingSummary("G9",sheet,type="heading")
        
        
#         sheet["G10"] = fb_video_data['VIEWS']
#         StylingSummary("G10",sheet)

#         sheet["G11"] = sc_video_data['VIEWS']
#         StylingSummary("G11",sheet)

#         # sheet["G12"] = ig_video_data['VIEWS']
#         sheet["G12"] = ig_video_data['VIEWS']
#         StylingSummary("G12",sheet)

#         sheet["G13"] = twitter_video_data['VIEWS']
#         StylingSummary("G13",sheet)

#         sheet["G14"] = dv360_data['VIEWS']
#         StylingSummary("G14",sheet)

#         sheet["G15"] = summary_data["VIEWS"]
#         StylingSummary("G15",sheet)

#         sheet["H9"] = "MONTHLY KPI ACHIEVEMENT"
#         StylingSummary("H9",sheet,type="heading")

#         sheet["H10"] = fb_video_data['MONTHLY_KPI_ACHIEVEMENT']
#         StylingSummary("H10",sheet)

#         sheet["H11"] = sc_video_data['MONTHLY_KPI_ACHIEVEMENT']
#         StylingSummary("H11",sheet)

#         # sheetH"G12"] = ig_video_data['MONTHLY_KPI_ACHIEVEMENT']
#         sheet["H12"] = ig_video_data['MONTHLY_KPI_ACHIEVEMENT']
#         StylingSummary("H12",sheet)

#         sheet["H13"] = twitter_video_data['MONTHLY_KPI_ACHIEVEMENT']
#         StylingSummary("H13",sheet)

#         sheet["H14"] = dv360_data['MONTHLY_KPI_ACHIEVEMENT']
#         StylingSummary("H14",sheet)
        
#         sheet["H15"] = summary_data["MONTHLY_KPI_ACHIEVEMENT"]
#         StylingSummary("H15",sheet)

#         sheet['I9'] = "CLICKS"
#         StylingSummary("I9",sheet,type="heading")

#         sheet["I10"] = fb_video_data['CLICKS']
#         StylingSummary("I10",sheet)

#         sheet["I11"] = sc_video_data['CLICKS']
#         StylingSummary("I11",sheet)
#         # sheetI"G12"] = ig_video_data['CLICKS']
#         sheet["I12"] = ig_video_data['CLICKS']
#         StylingSummary("I12",sheet)

#         sheet["I13"] = twitter_video_data['CLICKS']
#         StylingSummary("I13",sheet)

#         sheet["I14"] = dv360_data['CLICKS']
#         StylingSummary("I14",sheet)

#         sheet["I15"] = summary_data["CLICKS"]
#         StylingSummary("I15",sheet)

#         sheet["J9"] = "CTR"
#         StylingSummary("J9",sheet,type="heading")

#         sheet["J10"] = fb_video_data['CTR']
#         StylingSummary("J10",sheet)

#         sheet["J11"] = sc_video_data['CTR']
#         StylingSummary("J11",sheet)

#         # sheetJ"G12"] = ig_video_data['CTR']
#         sheet["J12"] = ig_video_data['CTR']
#         StylingSummary("J12",sheet)
#         sheet["J13"] = twitter_video_data['CTR']
#         StylingSummary("J13",sheet)
#         sheet["J14"] = dv360_data['CTR']
#         StylingSummary("J14",sheet)
#         sheet["J15"] = summary_data["CTR"]
#         StylingSummary("J15",sheet)

#         sheet["K9"] = "BUDGET"
#         StylingSummary("K9",sheet,type="heading")

#         sheet["K10"] = fb_video_data['BUDGET']
#         StylingSummary("K10",sheet)

#         sheet["K11"] = sc_video_data['BUDGET']
#         StylingSummary("K11",sheet)
#         # sheetK"G12"] = ig_video_data['BUDGET']
#         sheet["K12"] = ig_video_data['BUDGET']
#         StylingSummary("K12",sheet)
#         sheet["K13"] = twitter_video_data['BUDGET']
#         StylingSummary("K13",sheet)
#         sheet["K14"] = dv360_data['BUDGET']
#         StylingSummary("K14",sheet)
#         sheet["K15"] = summary_data["BUDGET"]
#         StylingSummary("K15",sheet)

#         sheet["L9"] = "BUDGET SPENT"
#         StylingSummary("L9",sheet,type="heading")
#         sheet["L10"] = fb_video_data['SPENT_BUDGET']
#         StylingSummary("L10",sheet)
#         sheet["L11"] = sc_video_data['SPENT_BUDGET']
#         StylingSummary("L11",sheet)
#         # sheetL"G12"] = ig_video_data['SPENT_BUDGET']
#         sheet["L12"] = ig_video_data['SPENT_BUDGET']
#         StylingSummary("L12",sheet)
#         sheet["L13"] = twitter_video_data['SPENT_BUDGET']
#         StylingSummary("L13",sheet)
#         sheet["L14"] = dv360_data['SPENT_BUDGET']
#         StylingSummary("L14",sheet)
#         sheet["L15"] = summary_data["SPENT_BUDGET"]
#         StylingSummary("L15",sheet)

#         default_template.save(filename+".xlsx")

#         serialized_workbook = pickle.dumps(default_template)
#         uploading_file = media_files(filename,serialized_workbook,current_user.username)
#         db.session.add(uploading_file)
#         db.session.commit()
#         # workbook.save("summary.xlsx")
#         # print("Runn")
#         # os.remove(filename+".xlsx")
#     except:
#         pass

#     # Save the file
#     # workbook.save("weekly_report.xlsx")



#     return jsonify({"message":"Uploaded successfully"}),200



# def StylingSummary(loc,sheet,type="data"):
#     thin_border = Border(left=Side(style='thin'), 
#     right=Side(style='thin'), 
#     top=Side(style='thin'), 
#     bottom=Side(style='thin'))

#     sheet[loc].font = Font(size=10, bold=False)
#     sheet[loc].font = sheet[loc].font.copy(name='맑은 고딕')
#     sheet[loc].alignment = sheet[loc].alignment.copy(horizontal="right") 
#     sheet[loc].border = thin_border

#     if type == "heading":
#         sheet[loc].font = Font(size=10,bold=True)
#         sheet[loc].font = sheet[loc].font.copy(name='맑은 고딕')
#         sheet[loc].alignment = sheet[loc].alignment.copy(wrap_text=True)
       
#         # sheet[loc].font = sheet[loc].font.copy(name='맑은 고딕')
#         # fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type = "solid")
#         sheet[loc].fill = PatternFill("solid", fgColor="DDDDDD") 
#         sheet[loc].alignment = sheet[loc].alignment.copy(horizontal="right") 

# def DataForSummary(filename,type):
#         res = db.engine.execute(f"SELECT * FROM `{filename}-weekly_total` WHERE `type` = '{type}';")
#         data = {

#         }
#         for dt in res:
#             if type == "trueview" or type == "gdn" or type == "dv360":
#                 fb_3_view = (dt['TrueView: Views'])
#             else:
#                 fb_3_view = (dt['3 Views'])
            
#             try:
#                 fb_monthly_view_kpi = dt['3 Views'] / dt['MVKPI']
#             except:
#                 fb_monthly_view_kpi = dt['View 100%'] / dt['MVKPI']
                
#             fb_clicks = dt['Clicks']
#             fb_ctr = dt['CTR']
#             fb_budget = dt['Budget']
#             fb_budget_spent = dt['SPENT_BUDGET']
#             # print(fb_budget_spent)
         
#             data["VIEWS"] = fb_3_view
           
#             data["MONTHLY_KPI_ACHIEVEMENT"] = str(round(fb_monthly_view_kpi*100))+ str("%")
#             data["CLICKS"] = fb_clicks
#             data["CTR"] = fb_ctr
#             data["BUDGET"] = fb_budget
#             data["SPENT_BUDGET"] = fb_budget_spent
#         return data

# def DataForSummaryTable():
#         res = db.engine.execute(f"SELECT * FROM `summary`;")
#         data = {

#         }
#         for dt in res:
          
#             fb_3_view = (dt['3 Views'])
        

#             fb_monthly_view_kpi = dt['3 Views'] / dt['MVKPI_Total']
           
                
#             fb_clicks = dt['Clicks']
#             fb_ctr = dt['CTR']
#             fb_budget = dt['Budget']
#             fb_budget_spent = dt['SPENT_BUDGET']
#             # print(fb_budget_spent)
         
#             data["VIEWS"] = fb_3_view
           
#             data["MONTHLY_KPI_ACHIEVEMENT"] = str(round(fb_monthly_view_kpi*100)) + str("%")
#             data["CLICKS"] = fb_clicks
#             data["CTR"] = fb_ctr
#             data["BUDGET"] = fb_budget
#             data["SPENT_BUDGET"] = fb_budget_spent
#             print(data)
#         return data



# @app.route("/fetch_files",methods=["GET"])
# @cross_origin()
# @token_required
# def fetch_files(current_user):
#     # Try catch if fail to connect with database 
#     res= db.session.query(media_files).filter_by(username=current_user.username).all()
#     db.session.close()

#     if res is None:
#         return jsonify({
#             "data": {},
#             "status": 200

#         },200)
#     data = []
#     for results in res:
#         data.append({"id":results.id,"value" : results.filename})
    
#     return jsonify({
#         "data":data,
#         "status" : 200
#     },200)
    





# @app.route("/download",methods=["POST"])
# @cross_origin()
# @token_required
# def download(current_user):
#         if request.method=="POST":
#             file_id = request.form['file_id']
#             ##print(file_id)
#             res = db.session.query(media_files).filter_by(id = file_id,username = current_user.username).first()
#             db.session.close()

#             if res is None:
#                 return jsonify({
#                     "message": "Data is deleted try another file",
#                     "status": 400
#                 },400)
            
#             workbook = pickle.loads(res.file)
#             filename_ext = str(res.filename)+".xlsx"
#             workbook.save(filename_ext)
#             return send_file(filename_ext),200

#     # Sending file issue here 
#     # Check if you can send  the file or not 



# @app.route("/fetch_weekly",methods=["POST"])
# @cross_origin()
# def fetch_weekly():
#         if request.method=="POST":
#             file_id = request.form['file_id']
#             ##print(file_id)
#             res = db.session.query(weekly_record).filter_by(file_id = file_id).all()
#             db.session.close()

#             if res is None:
#                 return jsonify({
#                     "weekly" : [],
#                     "status": 201
#                 },201)
            
#             weekly= []
#             # print(res)
#             for result in res:
#                 temp = {
#                     "x": str(result.week_end.month)+"/"+str(result.week_end.day),
#                     "date": result.week_end.date(),
#                     "impressions": int(result.impressions),
#                     "clicks": int(result.clicks),
#                     "ctr": result.CTR,
#                     "week_start": result.week_start,
#                     "trueview" : result.trueview,
#                     "Spendings" : result.Spendings,
#                     "Percent_25_views" : result.Percent_25_views,
#                     "Percent_50_views" : result.Percent_50_views,
#                     "Percent_75_views" : result.Percent_75_views,
#                     "Percent_100_views" : result.Percent_100_views,
#                     "Budget_Spent" : result.Budget_Spent,
#                     "Daily_View_KPI" : result.Daily_View_KPI,
#                     "CPV_TrueView" : result.CPV_TrueView,
#                     "CPV_Complete" : result.CPV_Complete,
#                     "Daily_KPI_Achievement" : result.Daily_KPI_Achievement,
#                     "Three_Views" : result.Three_Views,
#                     "CPV_Three" : result.CPV_Three,
#                 }
#                 weekly.append(temp)
            
#             return jsonify({
#                 "weekly": weekly
#             }),200

# @app.route("/fetch_annual",methods=["POST"])
# @cross_origin()
# def fetch_annual():
#         if request.method=="POST":
#             file_id = request.form['file_id']
#             ##print(file_id)
#             res = db.session.query(annual_record).filter_by(file_id = file_id).all()
#             db.session.close()

#             if res is None:
#                 return jsonify({
#                     "annual" : [],
#                     "status": 201
#                 },201)
            
#             annual= []
#             # print(res)
#             for result in res:
#                 temp = {
#                     "x": str(result.day_end.month)+"/"+str(result.day_end.day),
#                     "date": result.day_end.date(),
#                     "impressions": int(result.impressions),
#                     "clicks": int(result.clicks),
#                     "ctr": result.CTR,
#                     "day_start": result.day_start,
#                     "trueview" : result.trueview,
#                     "Spendings" : result.Spendings,
#                     "Percent_25_views" : result.Percent_25_views,
#                     "Percent_50_views" : result.Percent_50_views,
#                     "Percent_75_views" : result.Percent_75_views,
#                     "Percent_100_views" : result.Percent_100_views,
#                     "Budget_Spent" : result.Budget_Spent,
#                     "Daily_View_KPI" : result.Daily_View_KPI,
#                     "CPV_TrueView" : result.CPV_TrueView,
#                     "CPV_Complete" : result.CPV_Complete,
#                     "Daily_KPI_Achievement" : result.Daily_KPI_Achievement,
#                     "Three_Views" : result.Three_Views,
#                     "CPV_Three" : result.CPV_Three,
#                 }
#                 annual.append(temp)
            
#             return jsonify({
#                 "annual": annual
#             }),200





@app.route("/latest",methods=["POST"])
@cross_origin()
def latest():


    data_files = request.files
    # Fetching files from the request 
    data = request.form
    # Fetching data from request 


    total_raw_files = data["total_raw"]
    # Storing total raw data 

    raw = []
    # storing the data in the array for different media 
    
    # Total Raw Files for the loop 
    for i in range(1,int(total_raw_files)+1):
    # Looping through each raw file 

        label = "raw-"+str(i)+"-total_media"
        total_media = data[label]
        # Then Loop Total Media for each raw file


        # Data arr for each raw file 
        data_raw = []

        media = []

        for j in range(1,int(total_media)+1):
            # Storing each data in the dictionary and appending to array 

            label_media = "raw-"+str(i)+"-media-"+str(j)
            media_name = label_media + "-name"
            media_budget = label_media + "-total-budget"
            media_monthly_view_kpi = label_media + "-monthly-view-kpi"

            total_identifiers = data[label_media+"-total-identifiers"]
            
            identifiers = []

            for k in range(1,int(total_identifiers)+1):
                label_identifier = "raw-"+str(i)+"-media-"+str(j)+"-identifier-"+str(k)

                identifier_name = label_identifier + "-name"
                identifier_type = label_identifier+"-type"

                identifier_columns = []
                if data[identifier_type] == "multiple":
                    # If its multiple identifier then use this

                    for z in range(1,3):
                        column_name = data[label_identifier+"-column-"+str(z)]
                        column_identifier = data[label_identifier+"-column-"+str(z)+"-value"]
                        column_info = {
                             column_name: column_identifier
                        }
                        identifier_columns.append(column_info)
                        
                
                else:
                    # Else if its only single identifier then use this 

                    column_name = data[label_identifier+"-column-"+str(1)]
                    column_identifier = data[label_identifier+"-column-"+str(1)+"-value"]
                    column_info = {
                             column_name: column_identifier
                    }
                    identifier_columns.append(column_info)
                
                identifiers_info = {
                     "name": data[identifier_name],
                     "type": data[identifier_type],
                     "columns": identifier_columns
                }
                identifiers.append(identifiers_info)


                
            media_info = {
                "name":data[media_name],
                "budget": data[media_budget],
                "monthly_view_kpi":data[media_monthly_view_kpi],
                "total_identifiers": total_identifiers,
                "identifiers" : identifiers
            }
            media.append(media_info)



        

        file = data_files["raw-"+str(i)] 
        
       
        
        raw_info = {
            "raw": file,
            "total_media": total_media,
            "media": media
        }  

        raw.append(raw_info)  
        # This is the raw data with all the fields and information 
        

    print(raw)
    TotalBudgetForDaily = 0 
    TotalMonthlyViewKPI = 0

    list_of_dataframes = []
    # List to store the dataframes 
    deleting_table = False
    
    my_dir = os.path.dirname(__file__)
    pickle_file_path = os.path.join(my_dir, 'summary_testing.xlsx')
    with pd.ExcelWriter(pickle_file_path) as writer:
        list_of_summary_medias_with_data = [] 
        for dt in raw: 
        # Loop for going through each object in the dictionary to fetch the files and then all other information     
            file_data = pd.read_excel(dt["raw"])
            # reading excel sheet and storing the dataframe in the file_data variable 
        
            # Calculate the output according to the dataframe as before we were doing it 
            
            
            
            for x in range(0,int(dt["total_media"])):
            # Loop for going through each of the media and taking the data 

                TotalBudgetForDaily += int(dt["media"][x]["budget"])
                # Adding all the budgets for daily 
                TotalMonthlyViewKPI += int(dt["media"][x]["monthly_view_kpi"])
                # Adding all the monthly view kpi for daily 
                
                for y in range(0,int(dt["media"][x]["total_identifiers"])):
                # Loop for going through each of the identifiers as for now it is 2 

                    label_media_identifier = dt["media"][x]["name"]+"_"+dt["media"][x]["identifiers"][y]["name"]
                    # Getting the name of the identifier 

                    TotalBudget = dt["media"][x]["budget"]
                    # Getting the Budget of the identifier 
                    
                    MonthlyViewKPI = dt["media"][x]["monthly_view_kpi"]
                    # Getting the monthly view kpi for each identifier 

                    key_1 = list(dt["media"][x]["identifiers"][y]["columns"][0].keys())[0]
                    # Getting the first key as it is necessary in both cases either it is single or multiple 

                    value_1 = dt["media"][x]["identifiers"][y]["columns"][0][key_1]
                    # Getting the first key value and storing in the variable as it is necessary in both cases either it is single or multiple 
                    
                    key_2 = None
                    value_2 = None
                    # Setting key 2 and value 2  to none for now so if there is no key 2 then we can pass none 


                    if dt["media"][x]["identifiers"][y]["type"] == "multiple":
                    # If it is multiple then we can save the key 2 and value 2 values in the variable and replace none to that value 

                            key_2 = list(dt["media"][x]["identifiers"][y]["columns"][1].keys())[0]
                            # key 2 value is stored in the key 2 variable 

                            value_2 = dt["media"][x]["identifiers"][y]["columns"][1][key_2]
                            # value 2 value is stored in the value 2 variable 

                    # print(key_2,":",value_2)
                    # print(file_data[key_2].str.contains(value_2))
                    # Keys and values and full data from the raw file will be passed in the function to get the dataframe sorted into separate types  
                    # Find out the index column from the file instead of hard coding 
                    
                    index_col = file_data.columns[0]
                    # Storing the date column either it is Starting date or Date or something else 
                    # print(file_data)
                    dFrame = SeparatingDataframesForDifferentCategories(file_data=file_data,key_1=key_1,value_1=value_1,key_2=key_2,value_2=value_2, type = label_media_identifier,index_column=index_col,budget=TotalBudget,monthly_view_kpi=MonthlyViewKPI)
                    # Calling the function to set each type of dataframe like facebook or ig or sc or twitter and setting it to the dictionary object and naming it to the label of that  

                    dFrame_info = {
                        "name": label_media_identifier,
                        "data": dFrame
                    }
                    # print(label_media_identifier)
                    # Dataframe info is stored in this object 

                    # Weekly for each dataframe 
                    weekly_data_for_each_dFrame = CalculatingWeeklyForDailyTotalDataFrame(dFrame)
                    weekly_data_for_each_dFrame.index.name = "Weeks"
                    weekly_data_for_each_dFrame.index +=1


                    annual_total = weekly_data_for_each_dFrame.copy()
                    
                
                    annual_total.drop(weekly_data_for_each_dFrame.index, inplace=True)
                    # print(gdn_df_daily)
                    for i in range(0, 1):
                        row = weekly_data_for_each_dFrame.iloc[i:len(weekly_data_for_each_dFrame)].select_dtypes(include=['int64','double']).sum()
                        row['CTR'] = row['Clicks'] / row['Impressions']
                        # print(totalBudget)
                        row['Budget'] = round(row['Budget'],2)
                        row['SPENT_BUDGET'] = round((row['Budget']/TotalBudgetForDaily)*100)
                        row['CTR'] = round(row['CTR']*100,2)
                    
                        annual_total = annual_total.append(row, ignore_index=True)
                    
                    
                    annual_total = annual_total.assign(MVKPI_Total = MonthlyViewKPI)





                    try:
                        # weekly_data_for_each_dFrame.to_sql(label_media_identifier+"_weekly",con= db.engine,if_exists="replace",index=True,index_label="Weeks")
                        
                        annual_total = annual_total.assign(CPC = round((annual_total["Budget"] / annual_total["Clicks"]),2))

                        # annual_total.to_sql(label_media_identifier+"_final",con= db.engine,if_exists="replace",index=True)
                        list_of_summary_medias_with_data.append({
                            "label": label_media_identifier,
                            "data":annual_total,
                            "budget": TotalBudget,
                        })
                        # Data is stored in the excel file 
                        weekly_data_for_each_dFrame.to_excel(writer, sheet_name=label_media_identifier, index=True,index_label=index_col)
                        
                        annual_total.to_excel(writer, sheet_name=label_media_identifier,startrow=writer.sheets[label_media_identifier].max_row, index=True,header=False)
                        dFrame.to_excel(writer, sheet_name=label_media_identifier,startrow=writer.sheets[label_media_identifier].max_row+1, index=True)
                    except:
                        # db.engine.execute(f"DROP table IF EXISTS `{label_media_identifier}`")
                        label_media_identifier_weekly=label_media_identifier+"_weekly"
                        # db.engine.execute(f"DROP table IF EXISTS `{label_media_identifier_weekly}`")
                        label_media_identifier_final=label_media_identifier+"_final"
                        # db.engine.execute(f"DROP table IF EXISTS `{label_media_identifier_final}`")
                        deleting_table = True


                    list_of_dataframes.append(dFrame_info)
                    # Appending the dataframe in the list of dataframe

        # print(list_of_dataframes)
        # Total from raw data according to date or reporting starts
        temp = pd.DataFrame()
        # Empty dataframe for combining all the dataframes in the daily total tab 

        # Empty dataframe to append in all other dataframes 
        
        # From here we are trying to concat all the dataframes for the daily total
        
        for dt in list_of_dataframes:
            # Loop to go through each dataframe and concat it to the previous dataframe 

            if deleting_table == True:
                pass
                # db.engine.execute(f"DROP table IF EXISTS `{dt['name']}`")

            else:

                # Commenting weekly label                   (will use later)
                # weekly_label = dt["name"] + "-weekly"
                

                try:
                    if dt["data"].index.name != "Date":
                        dt["data"].index.name = "Date"
                    # dt["data"].to_sql(dt["name"],con=db.engine,if_exists="replace",index=True,index_label="Date")
                    
                    
                    # Commenting weekly dataframe (Will use it later) 
                    # weekly_dataframe = CalculatingWeeklyForDailyDataFrame(dt["data"],column_type=dt["name"])
                    # weekly_dataframe.to_sql(weekly_label,con=db.engine,if_exists="replace",index=True)


                    temp = pd.concat([temp, dt["data"]])
                    
                except:
                    # db.engine.execute(f"DROP table IF EXISTS `{dt['name']}`")
                    pass

        
        # print(temp)
        

                # Commenting drop table weekly label (Use it later)
                # db.engine.execute(f"DROP table IF EXISTS `{weekly_label}`")
        # print(temp)
        # print(temp)
        # For daily report tab and weekly for daily tab  
        # print(temp)
        
        # We are getting all the data together in one dataframe now we have to sum according to date 
        # To do that we have to use pivot table function and add sum function 
        # But first we have to create a variable in which we will store the columns 
        # Also we need a variable for storing the index 
        if deleting_table == True:
            label_daily_total = "daily_total"
            label_daily_total_weeks = "daily_total_weekly"
            label_summary_total = "summary_total"

            # db.engine.execute(f"DROP table IF EXISTS `{label_daily_total}`")
            # db.engine.execute(f"DROP table IF EXISTS `{label_daily_total_weeks}`")
            # db.engine.execute(f"DROP table IF EXISTS `{label_summary_total}`")

        else:

            index_name = "Date"
            # print(temp)
            column_values = temp.columns.ravel()
            # print(column_values)
            daily_total=  pd.pivot_table(temp, index=[index_name],values=column_values,aggfunc='sum')
            # print(daily_total)


            # print("====================================================")
            # print(daily_total["Impressions"])
            # print("====================================================")

            # Now we need to calculate other columns of the daily total which has formulas 
            # daily_total = SettingRemainingColumnsOfDailyTotalDataFrame(daily_total,TotalBudgetForDaily,TotalMonthlyViewKPI)

            label_daily_total = "daily_total"
            # daily_total.to_sql(label_daily_total,con = db.engine,if_exists="replace",index=True)
            # daily_total.to_excel(writer, sheet_name=label_daily_total, index=True)
            # print(daily_total)
            daily_total_weekly_dataframe = CalculatingWeeklyForDailyTotalDataFrame(daily_total)
            daily_total_weekly_dataframe.index.name = "Weeks"
            daily_total_weekly_dataframe.index +=1
            # daily_total_weekly_dataframe.to_sql("daily_total_weekly",con= db.engine,if_exists="replace",index=True,index_label="Weeks")

           
            # daily_total_weekly_dataframe.to_excel(writer, sheet_name="daily_total_weekly", index=True)



            # Calculating the total for daily total weekly dataframe 
            #  
            # daily_total_weekly_dataframe.index = 0
            summary_total =  daily_total_weekly_dataframe.sum()
            # Add formulas for formula columns 
            

            # summary_total.to_sql("summary_total",con= db.engine, if_exists="replace",index=False)
            
            
            
            # print(summary_total)


            # 
            # *********** 
            # Summary logic =============================================================
            # 

            summary = daily_total_weekly_dataframe.copy()
                        
                    
            summary.drop(summary.index, inplace=True)
            # print(gdn_df_daily)
            for i in range(0, 1):
                row = daily_total_weekly_dataframe.iloc[i:len(daily_total_weekly_dataframe)].select_dtypes(include=['int64','double']).sum()
                row['CTR'] = row['Clicks'] / row['Impressions']
                # print(totalBudget)
                row['Budget'] = round(row['Budget'],2)
                row['SPENT_BUDGET'] = round((row['Budget']/TotalBudgetForDaily)*100)
                row['CTR'] = round(row['CTR']*100,2)
            
                summary = summary.append(row, ignore_index=True)
            
            
            summary = summary.assign(MVKPI_Total = TotalMonthlyViewKPI)
            summary = summary.assign(CPC = round((summary["Budget"] / summary["Clicks"]),2))
            # summary.to_sql("summary_total",con=db.engine,if_exists="replace",index=False)
            list_of_summary_medias_with_data.append({
                "label": "Total",
                "data":summary,
                "budget": int(TotalBudgetForDaily),

            })
            daily_total_weekly_dataframe.to_excel(writer, sheet_name=label_daily_total, index=True)
            summary.to_excel(writer, sheet_name=label_daily_total,startrow=writer.sheets[label_daily_total].max_row, index=True,header=False)
            daily_total.to_excel(writer, sheet_name=label_daily_total,startrow=writer.sheets[label_daily_total].max_row+1, index=True)
            # summary.to_excel(writer, sheet_name="summary_total", index=False)

        # summary.to_excel(writer, sheet_name='summary', index=False)
    # summary.to_excel(writer,con=db.engine,if_exist="replace",index=False)
    
    # new_df.set_index('Date', inplace=True)
    # daily_total_weekly_dataframe.to_excel(writer, sheet_name='daily_report', index=True,index_label="Date")
    # writer = StylingSheets(writer, "daily_report")
    #=============== Summary logic ends ============================================================
        count_start = 0
        
        df2 = pd.DataFrame() 
        for dt in list_of_summary_medias_with_data:
                print(dt["data"]["Budget"].values[0])
                Media_Column = dt["label"]
                Budget_Column = dt["budget"]
                Spent_Column = dt["data"]["Budget"].values[0]
                SPENT_BUDGET_Column = dt["data"]["SPENT_BUDGET"].values[0]
                Clicks_Column = dt["data"]["Clicks"].values[0]
                CTR_Column = dt["data"]["CTR"].values[0]
                CPC = dt["data"]["CPC"].values[0] 
                df = pd.DataFrame([[Media_Column,Budget_Column,Spent_Column,SPENT_BUDGET_Column,Clicks_Column,CTR_Column,CPC]],columns=['Media','Budget','Spent','Budget Spent','Clicks','CTR','CPC'])
                df2 = df2.append(df)

        # print(df2)
        df2.to_excel(writer, sheet_name="Summary",index=False)

    # Now next is weekly _total so we can move on to summary 

    # daily_total_weekly_total =  pd.pivot_table(daily_total_weekly_dataframe,values=column_values,aggfunc = 'sum')


    # print(list_of_dataframes)
    
    return send_file(writer),200
    # return jsonify({
    #      "message": "Got data and stored in object successfully",
    #      "status": 200
    # },200)
    # Returning the request to show that the data is receiving or not 

def SeparatingDataframesForDifferentCategories(file_data,key_1,value_1,key_2,value_2,type,index_column,budget,monthly_view_kpi):
    
    # getting dataframe to compare for first column to compare 
    data_frame = file_data[file_data[key_1].str.contains(value_1)]
    # So we are creating a dataframe and filtering the value 1 is present in the key 1 dataframe column 

    # print(data_frame[key_2].str.contains(value_2))
    if key_2 != None and value_2 != None:
        # If key 2 and value 2 is not none then we can filter 2nd column according to that and get the dataframe 
    # getting dataframe to compare for second column to compare 
        data_frame = data_frame[data_frame[key_2].str.contains(value_2)]    
        # Setting dataframe ...
    # print(data_frame)
    column_values = data_frame.columns.ravel()
    # Column names for the dataframe will be stored in the column values 

    # print(column_values)
    data_frame[index_column] = pd.to_datetime(data_frame[index_column])
    # Converting the index column to date time so we can have a same format for all dates and not mixed like text and date .........

    # data_frame[index_column] = data_frame[index_column].astype("datetime")
    data_frame = pd.pivot_table(data_frame, index=[index_column],values=column_values,aggfunc='sum')  
    # We are summing the dataframe and storing the dataframe with sum of similar dates record 
    # print(data_frame)
    data_frame.index.name = "Date"
    data_frame = SettingRemainingColumnsOfDailyTotalDataFrame(df=data_frame,budget=budget,monthly_view_kpi = monthly_view_kpi)
    # Now we want to calculate the remaining columns here so we are passing dataframe and budget and monthly view kpi in the function .....
    
    data_frame = data_frame.assign(type = type)
 
    # returning the dataframe with comparing two columns where its same 
    return data_frame


def SettingRemainingColumnsOfDailyTotalDataFrame(df,budget,monthly_view_kpi):
    df['Budget'] = (round(df['Budget'],2)).replace([np.inf, -np.inf,np.nan], 0)
    df = df.assign(DAILY_VIEW_KPI = (round(float(monthly_view_kpi) / 30)))
    df = df.assign(DAILY_KPI_ACHIEVEMENT = (round((df['View 100%'] / df['DAILY_VIEW_KPI'])*100)).replace([np.inf, -np.inf], 0))
    df = df.assign(SPENT_BUDGET = (round((df['Budget'].cumsum()/int(budget))*100)).replace([np.inf, -np.inf], 0))
    df = df.assign(CPV_COMPLETE = (round(df["Budget"]/df["View 100%"],4)).replace([np.inf, -np.inf], 0))
    df = df.assign(CTR = (round(((df["Clicks"]/df["Impressions"])*100),2).replace([np.inf, -np.inf], 0)))
    try:
        df = df.assign(CPV_TRUEVIEW = (round(df["Budget"]/df["ThruPlays"],4)).replace([np.inf, -np.inf], 0))
    except:
        df = df.assign(CPV_TRUEVIEW = (round(df["Budget"]/df["TrueView: Views"],4)).replace([np.inf, -np.inf], 0))
    
        
    return df

def CalculatingWeeklyForDailyDataFrame(df_daily,column_type):
    date_rng = pd.date_range(start=df_daily.index.min(), end=df_daily.index.max(), freq='D')
    # print("Fine")
    # Fetching the start and end date so we can loop through it for weeks 
    df_weekly = df_daily.copy()
    # Making a copy of daily table so we can use the same columns for weekly as well
    # print("Still fine")
    df_weekly.drop(df_weekly.index, inplace=True)
    # We are dropping all the data so we can add new data inside weekly 
    # print("still still fine")

    # print(gdn_df_daily)
    # Starting the loop for weeks 
    for i in range(0, len(date_rng), 7):
        try:
            row = df_daily.iloc[i:i+7].select_dtypes(include=['int64','double']).sum()
        except:
            print("not fine 1")
        # Doing sum for the row from 1 - 7 days and storing in row 
        try:
            row['Budget'] = round(row['Budget'],2)
        # Changing the Format of the budget 
        except:
            print("not fine 2")
        try:
            row['CTR']=round(((row['Clicks']/row['Impressions'])*100),2)
        except:
            print("not fine 3")
        # Changing the CTR for the week 
        # row['CTR'] = round(row['CTR'],2)
        # Changing the format for CTR

        try:
            row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
        except:
            print("not fine 4")
        
        try:
            try:
                if row["ThruPlays"] == 0:
                    row['CPV_TRUEVIEW'] = 0
                else:
                    row['CPV_TRUEVIEW'] = round(row["Budget"]/row["ThruPlays"],4)

            except:
                if row["TrueView: Views"] == 0:
                    row['CPV_TRUEVIEW'] = 0
                else:
                    row['CPV_TRUEVIEW'] = round(row["Budget"]/row["TrueView: Views"],4)
        except:
            print("not fine 5")
        # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
        if row["View 100%"] == 0:
            row['CPV_COMPLETE'] = 0
        else:
            row['CPV_COMPLETE'] = round(row["Budget"]/row["View 100%"],4)

        if i+6 > len(date_rng):
          
            row["SPENT_BUDGET"] = df_daily["SPENT_BUDGET"][len(date_rng)-1]
        else : 
            # row['Budget'] = round(row['Budget'],2)
            row["SPENT_BUDGET"] = df_daily["SPENT_BUDGET"][i+6]
            # row['CTR'] = round(row['CTR'],2)
            # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
            # row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
            row["DAILY_KPI_ACHIEVEMENT"] = round(row["View 100%"]/ row["DAILY_VIEW_KPI"]*100)
            # if type == "fb_video" or type == "fb_image" or type=="ig_image" or type=="ig_video":
            # else:
            #     row["DAILY_KPI_ACHIEVEMENT"] = round(row["View 100%"]/ row["DAILY_VIEW_KPI"]*100)
        row["type"] = column_type 
        df_weekly = df_weekly.append(row, ignore_index=True)
    # print("absolutely fine")
    # del df_weekly['type']
    return df_weekly 


def CalculatingWeeklyForDailyTotalDataFrame(df_daily):
    date_rng = pd.date_range(start=df_daily.index.min(), end=df_daily.index.max(), freq='D')
    # print("Fine")
    # Fetching the start and end date so we can loop through it for weeks 
    df_weekly = df_daily.copy()
    # Making a copy of daily table so we can use the same columns for weekly as well
    # print("Still fine")
    df_weekly.drop(df_weekly.index, inplace=True)
    # We are dropping all the data so we can add new data inside weekly 
    # print("still still fine")

    # print(gdn_df_daily)
    # Starting the loop for weeks 
    for i in range(0, len(date_rng), 7):
        try:
            row = df_daily.iloc[i:i+7].select_dtypes(include=['int64','double']).sum()
        except:
            print("not fine 1")
        # Doing sum for the row from 1 - 7 days and storing in row 
        try:
            row['Budget'] = round(row['Budget'],2)
        # Changing the Format of the budget 
        except:
            print("not fine 2")
        try:
            row['CTR']=round(((row['Clicks']/row['Impressions'])*100),2)
        except:
            print("not fine 3")
        # Changing the CTR for the week 
        # row['CTR'] = round(row['CTR'],2)
        # Changing the format for CTR

        try:
            row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
        except:
            print("not fine 4")
        
        try:
            try:
                if row["ThruPlays"] == 0:
                    row['CPV_TRUEVIEW'] = 0
                else:
                    row['CPV_TRUEVIEW'] = round(row["Budget"]/row["ThruPlays"],4)

            except:
                if row["TrueView: Views"] == 0:
                    row['CPV_TRUEVIEW'] = 0
                else:
                    row['CPV_TRUEVIEW'] = round(row["Budget"]/row["TrueView: Views"],4)
        except:
            print("not fine 5")
        # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
        if row["View 100%"] == 0:
            row['CPV_COMPLETE'] = 0
        else:
            row['CPV_COMPLETE'] = round(row["Budget"]/row["View 100%"],4)

        print("Position of index:",i+6)
        print(len(date_rng))
        if i+6 >= len(date_rng):
            # print(row["SPENT_BUDGET"])
            row["SPENT_BUDGET"] = df_daily["SPENT_BUDGET"][len(date_rng)-1]
        else : 
            # row['Budget'] = round(row['Budget'],2)
            row["SPENT_BUDGET"] = df_daily["SPENT_BUDGET"][i+6]
            # row['CTR'] = round(row['CTR'],2)
            # row['CPV_COMPLETE'] = round(row['CPV_COMPLETE'],4)
            # row['CPV_TRUEVIEW'] = round(row['CPV_TRUEVIEW'],4)
            row["DAILY_KPI_ACHIEVEMENT"] = round(row["View 100%"]/ row["DAILY_VIEW_KPI"]*100)
            # if type == "fb_video" or type == "fb_image" or type=="ig_image" or type=="ig_video":
            # else:
            #     row["DAILY_KPI_ACHIEVEMENT"] = round(row["View 100%"]/ row["DAILY_VIEW_KPI"]*100)
        
        df_weekly = df_weekly.append(row, ignore_index=True)
    # print("absolutely fine")
    # del df_weekly['type']
    return df_weekly 

@app.route("/",methods=["GET"])
def form_page():
    return render_template("frontend/index.html")

if __name__ == "__main__":
	app.run(debug = True)




